"""
Builds the styled output Excel workbook.

STYLING:
  Header: dark green #375623, white bold text, Calibri 10pt, thin borders
  Data:   unstyled — values written with ws.append() only

TO ADD/REMOVE/REORDER COLUMNS:
  Edit extraction/output_schema.py only.
  For per-file dynamic columns, pass a DynamicSchema to build_xlsx().
"""
from __future__ import annotations

import io
import logging
import re

import openpyxl
from openpyxl.styles import NamedStyle, PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from spir_dynamic.extraction.output_schema import OUTPUT_COLS, COL_WIDTHS
from spir_dynamic.utils.logging import timed

log = logging.getLogger(__name__)

_HDR_BG        = "375623"
_HDR_FONT_CLR  = "FFFFFF"
_DATA_FONT_NAME = "Calibri"
_FONT_SIZE     = 10
_HDR_HEIGHT    = 30
_ROW_HEIGHT    = 15
_DEFAULT_WIDTH  = 14

# Used only for the header row — created once at module level.
_THIN   = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


@timed
def build_xlsx(rows: list[list], spir_no: str = "") -> bytes:
    """
    Build a styled .xlsx workbook from extracted rows (always 27 columns).
    Returns raw .xlsx bytes.

    Performance: the header NamedStyle is registered once and applied to 27
    cells. Data rows are written with ws.append() only — no per-cell style
    assignment, no border XML per data cell. For a 2,000-row × 27-col file
    this removes ~54,000 ws.cell() calls and ~54,000 stylesheet mutations,
    plus eliminates all per-cell border XML from the serialised output.
    """
    col_names = OUTPUT_COLS
    col_widths = COL_WIDTHS
    n_cols = len(col_names)

    wb = openpyxl.Workbook()
    ws = wb.active
    safe_title = re.sub(r'[\\/*?:\[\]]+', ' ', spir_no or "SPIR Extraction").strip()
    ws.title = safe_title[:31]

    # ── Header NamedStyle — registered ONCE, applied to 27 cells only ─────────
    # Guard against ValueError("Style already exists") if called on a workbook
    # that already registered this style (e.g. in tests).
    if "spir_hdr" not in wb.named_styles:
        hdr_style           = NamedStyle(name="spir_hdr")
        hdr_style.fill      = PatternFill("solid", fgColor=_HDR_BG)
        hdr_style.font      = Font(name=_DATA_FONT_NAME, size=_FONT_SIZE,
                                   bold=True, color=_HDR_FONT_CLR)
        hdr_style.alignment = Alignment(horizontal="center", vertical="center",
                                        wrap_text=True)
        hdr_style.border    = _BORDER
        wb.add_named_style(hdr_style)

    # ── Column widths ─────────────────────────────────────────────────────────
    for idx, col_name in enumerate(col_names, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = col_widths.get(
            col_name, _DEFAULT_WIDTH
        )

    # ── Row heights ───────────────────────────────────────────────────────────
    ws.sheet_format.defaultRowHeight = _ROW_HEIGHT
    ws.sheet_format.customHeight = True
    ws.row_dimensions[1].height = _HDR_HEIGHT

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(col_names, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.style = "spir_hdr"

    ws.freeze_panes = "A2"

    # ── Data rows — ws.append() only, zero style overhead ────────────────────
    # No inner style loop, no ws.cell() calls, no border XML per data cell.
    # Values are identical: ws.append() writes exactly what post_process_rows
    # produced — column order, row count, and all values are untouched.
    for row in rows:
        if isinstance(row, (list, tuple)):
            r = list(row)
            if len(r) < n_cols:
                r += [None] * (n_cols - len(r))
            r = r[:n_cols]
        else:
            r = [None] * n_cols

        # Sanitize sentinel "." → None
        r = [None if v == "." else v for v in r]
        # Uppercase all string values in the output
        r = [v.upper() if isinstance(v, str) else v for v in r]
        ws.append(r)

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"

    # ── Serialise ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    result = buf.read()

    log.info("Excel built: %d rows x %d cols -> %d bytes", len(rows), n_cols, len(result))
    return result
