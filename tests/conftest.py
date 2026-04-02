"""
Test fixtures — helpers to create mock openpyxl workbooks for testing.
"""
from __future__ import annotations

import pytest
import openpyxl


@pytest.fixture
def make_workbook():
    """Factory fixture that creates a workbook from a list of sheet definitions."""

    def _make(sheets: dict[str, list[list]]) -> openpyxl.Workbook:
        """
        Create a workbook from sheet data.

        Args:
            sheets: {sheet_name: [[row1_values], [row2_values], ...]}

        Returns:
            openpyxl.Workbook with the specified data.
        """
        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        for sheet_name, rows in sheets.items():
            ws = wb.create_sheet(title=sheet_name)
            for r_idx, row_data in enumerate(rows, start=1):
                for c_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        return wb

    return _make


@pytest.fixture
def tabular_workbook(make_workbook):
    """A workbook with a standard tabular layout (tag column + data columns)."""
    return make_workbook({
        "Sheet1": [
            ["Tag No", "Description", "Quantity", "Unit Price", "Part Number", "UOM"],
            ["30-P-001", "Bearing assembly", 2, 150.00, "BRG-100", "EA"],
            ["30-P-001", "Seal kit", 4, 25.50, "SK-200", "SET"],
            ["30-P-002", "Impeller", 1, 800.00, "IMP-300", "EA"],
            ["30-P-002", "O-ring set", 6, 12.00, "OR-400", "SET"],
        ],
    })


@pytest.fixture
def columnar_workbook(make_workbook):
    """A workbook with a matrix/columnar layout (tags as column headers)."""
    return make_workbook({
        "SPIR Matrix": [
            # Row 1: Tag headers
            [None, None, None, "30-GV-146", "30-GV-171", "30-GV-169"],
            # Row 2: Manufacturer
            ["Manufacturer", None, None, "Flowserve", "Flowserve", "Flowserve"],
            # Row 3: Model
            ["Model", None, None, "Mark III", "Mark III", "Mark IV"],
            # Row 4-6: Other metadata rows
            [None, None, None, None, None, None],
            [None, None, None, None, None, None],
            [None, None, None, None, None, None],
            # Row 7: Data header
            ["Item Number", "Description", "Unit Price", "Qty", "Qty", "Qty"],
            # Row 8+: Data rows
            [1, "Bearing assembly", 150.00, 2, 2, 3],
            [2, "Seal kit", 25.50, 4, 4, 6],
            [3, "Impeller", 800.00, 1, 1, 1],
        ],
    })


@pytest.fixture
def multi_sheet_workbook(make_workbook):
    """A workbook with multiple data sheets including a continuation."""
    return make_workbook({
        "Main Data": [
            ["Tag No", "Description", "Quantity", "Unit Price", "Part Number"],
            ["30-P-001", "Bearing assembly", 2, 150.00, "BRG-100"],
            ["30-P-001", "Seal kit", 4, 25.50, "SK-200"],
        ],
        "Continuation Sheet": [
            ["Tag No", "Description", "Quantity", "Unit Price", "Part Number"],
            ["30-P-001", "Impeller", 1, 800.00, "IMP-300"],
            ["30-P-001", "O-ring set", 6, 12.00, "OR-400"],
        ],
        "Validation": [
            ["Dropdown Values"],
            ["EA"],
            ["SET"],
            ["KG"],
        ],
    })


@pytest.fixture
def annexure_workbook(make_workbook):
    """A workbook with annexure-style layout (tags as row headers)."""
    return make_workbook({
        "Annexure 1": [
            [None, "Description", "Quantity", "Part Number", "Unit Price"],
            ["30-P-001", "Bearing assembly", 2, "BRG-100", 150.00],
            ["30-P-002", "Seal kit", 4, "SK-200", 25.50],
            ["30-P-003", "Impeller", 1, "IMP-300", 800.00],
        ],
    })
