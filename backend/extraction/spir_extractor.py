"""
extraction/spir_extractor.py
─────────────────────────────
Qatar Energy SPIR matrix extractor — all verified formats.

OUTPUT COLUMNS (27 total):
  SPIR NO | TAG NO | EQPT MAKE | EQPT MODEL | EQPT SR NO | EQPT QTY |
  QUANTITY IDENTICAL PARTS FITTED | ITEM NUMBER | POSITION NUMBER |
  OLD MATERIAL NUMBER/SPF NUMBER | DESCRIPTION OF PARTS |
  NEW DESCRIPTION OF PARTS | DWG NO INCL POSN NO |
  MANUFACTURER PART NUMBER | MATERIAL SPECIFICATION |
  SUPPLIER OCM NAME | CURRENCY | UNIT PRICE | UNIT PRICE (QAR) |
  DELIVERY TIME IN WEEKS | MIN MAX STOCK LVLS QTY | UNIT OF MEASURE |
  SAP NUMBER | CLASSIFICATION OF PARTS | SPIR ERROR | SHEET | SPIR TYPE

SPIR TYPE values (exactly 4 allowed):
  1. Normal Operating Spares
  2. Initial Spare Parts
  3. Commissioning Spare Parts
  4. Life Cycle Spare Parts

OLD MATERIAL NUMBER rules:
  - Always exactly 18 characters
  - Base = numeric segments of SPIR_NO (after removing VEN- prefix and alpha codes)
  - Suffix = '-L' + zero-padded item number (within 18-char budget)
  - If budget exceeded: use '-{first_digit}L{rest}' format
  - If still exceeded: remove second hyphen from base and retry

POSITION NUMBER:
  - Always 4-digit zero-padded: item_num * 10 → '0010', '0020', etc.
  - For items > 999: str(item_num * 10).zfill(4)

EQPT MAKE:
  - Taken from MANUFACTURER field in the SPIR header (row 3, right-side label)
  - NOT the supplier name — the equipment manufacturer
"""
from __future__ import annotations
import re
import logging
from typing import Optional

import openpyxl

log = logging.getLogger(__name__)

# ── Output column order (27 columns) ─────────────────────────────────────────
OUTPUT_COLS = [
    'SPIR NO',
    'TAG NO',
    'EQPT MAKE',
    'EQPT MODEL',
    'EQPT SR NO',
    'EQPT QTY',
    'QUANTITY IDENTICAL PARTS FITTED',
    'ITEM NUMBER',
    'POSITION NUMBER',
    'OLD MATERIAL NUMBER/SPF NUMBER',
    'DESCRIPTION OF PARTS',
    'NEW DESCRIPTION OF PARTS',
    'DWG NO INCL POSN NO',
    'MANUFACTURER PART NUMBER',
    'MATERIAL SPECIFICATION',
    'SUPPLIER OCM NAME',
    'CURRENCY',
    'UNIT PRICE',
    'UNIT PRICE (QAR)',
    'DELIVERY TIME IN WEEKS',
    'MIN MAX STOCK LVLS QTY',
    'UNIT OF MEASURE',
    'SAP NUMBER',
    'CLASSIFICATION OF PARTS',
    'SPIR ERROR',
    'SHEET',
    'SPIR TYPE',
]

# ── SPIR TYPE normalisation ───────────────────────────────────────────────────
_SPIR_TYPES = {
    'normal': 'Normal Operating Spares',
    'initial': 'Initial Spare Parts',
    'commission': 'Commissioning Spare Parts',
    'life': 'Life Cycle Spare Parts',
}

def _normalise_spir_type(raw: str) -> str:
    """Normalise any SPIR type string to one of the 4 allowed values."""
    lower = raw.lower().strip()
    if 'commission' in lower:
        return 'Commissioning Spare Parts'
    if 'initial' in lower:
        return 'Initial Spare Parts'
    if 'life' in lower or 'lifecycle' in lower:
        return 'Life Cycle Spare Parts'
    return 'Normal Operating Spares'   # default

# ── Column label → field name (row 7 numeric labels) ─────────────────────────
_COL_LABEL_TO_FIELD = {
    "7":   "qty_identical",
    "8":   "desc",
    "9":   "dwg_no",
    "10a": "mfr_part_no",
    "10b": "supplier_part_no",
    "11a": "material_spec",
    "11b": "material_cert",
    "12a": "supplier_name",
    "12b": "supporting_docs",
    "17":  "currency",
    "18":  "delivery",
    "19":  "min_max",
    "20":  "uom",
    "21":  "sap_no",
    "22":  "classification",
}

_HEADER_FALLBACK = {
    "total no. of identical": "qty_identical",
    "identical parts":        "qty_identical",
    "description of parts":   "desc",
    "dwg no":                 "dwg_no",
    "posn no":                "dwg_no",
    "manufacturer part":      "mfr_part_no",
    "suppliers part":         "supplier_part_no",
    "material spec":          "material_spec",
    "material cert":          "material_cert",
    "supplier/ocm":           "supplier_name",
    "supplier ocm":           "supplier_name",
    "currency":               "currency",
    "unit price":             "unit_price",
    "delivery":               "delivery",
    "stock lvls":             "min_max",
    "unit of measure":        "uom",
    "unit  of measure":       "uom",
    "sap number":             "sap_no",
    "sap no":                 "sap_no",
    "classification":         "classification",
}


def _v(ws, r: int, c: int) -> str:
    val = ws.cell(r, c).value
    if val is None: return ""
    s = str(val).strip()
    return "" if s in (".", "_") else s


def _raw(ws, r: int, c: int):
    return ws.cell(r, c).value


def _is_tag(val: str) -> bool:
    if not val or len(val) > 80: return False
    skip = {"spare", "interchangeability", "record", "spir", "note",
            "authority", "required", "quantities", "authorised",
            "equipment", "tag no", "annexure", "refer", "qatar",
            "equip", "or tag", "or model", "mfr", "serial", "no.",
            "_", "spir number", "na", "n/a", "tba"}
    lower = val.lower().strip()
    if any(lower == w for w in skip): return False
    if any(w in lower for w in ["spare parts", "interchangeability", "record",
                                 "spir number", "authority", "required on site",
                                 "quantities in col"]): return False
    return True


def _expand_tags(raw: str) -> list[str]:
    raw = raw.strip()
    if not raw: return []
    m = re.match(r'^(\d+)\s+to\s+(\d+)$', raw, re.IGNORECASE)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        return [str(i) for i in range(a, b + 1)]
    if "/" in raw or "," in raw:
        parts = [p.strip() for p in re.split(r'[/,]', raw) if p.strip()]
        if len(parts) > 1:
            first = parts[0]
            result = [first]
            for part in parts[1:]:
                if re.match(r'^\d+$', part) and len(part) < len(first):
                    pm = re.match(r'^(.*\D)(\d+)$', first)
                    result.append(pm.group(1) + part if pm else part)
                else:
                    result.append(part)
            return result
    return [raw]


def _detect_spir_type(ws) -> str:
    """
    Detect the SPIR type from the sheet headers.
    Scans ALL rows 1-11 and collects all type mentions.
    If BOTH Initial and Normal are mentioned (e.g. col 17 = Initial, col 19 = Normal),
    we prefer 'Normal Operating Spares' as the sheet-level default because the
    per-item logic will override with 'Initial' for specific items when col 17 > 0.
    """
    found_types: list[str] = []
    for r in range(1, 12):
        for c in range(1, min(ws.max_column + 1, 30)):
            v = _v(ws, r, c).lower()
            if "normal operating" in v or "normal operation" in v:
                found_types.append("Normal Operating Spares")
            elif "initial spare" in v:
                found_types.append("Initial Spare Parts")
            elif "commissioning" in v:
                found_types.append("Commissioning Spare Parts")
            elif "life cycle" in v:
                found_types.append("Life Cycle Spare Parts")
    if not found_types:
        return "Normal Operating Spares"
    # If Normal appears anywhere → use it as default (Initial will be set per-item via flags)
    if "Normal Operating Spares" in found_types:
        return "Normal Operating Spares"
    return found_types[0]


def _detect_spir_no(wb) -> str:
    pattern = re.compile(
        r'(VEN-\d{4}-[A-Z0-9]+-\d+-\d+-[\w.\-]+|'
        r'\d{4}-[A-Z]{2}-\d{2}-\d{2}-\d{2}-\d{3}[\w.\-]*)',
        re.IGNORECASE)
    for sn in wb.sheetnames:
        if "validation" in sn.lower(): continue
        ws = wb[sn]
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(1, min(50, ws.max_column + 1)):
                v = str(ws.cell(r, c).value or "")
                m = pattern.search(v)
                if m: return m.group(1).strip()
    return ""


# ── OLD MATERIAL NUMBER builder ───────────────────────────────────────────────

def build_old_material_number(spir_no: str, item_num: int) -> str:
    """
    Build OLD MATERIAL NUMBER exactly 18 characters from SPIR_NO + item_num.

    Steps:
    1. Extract numeric base from SPIR_NO:
       remove 'VEN-' prefix, keep only segments starting with a digit
       e.g. VEN-4460-KAHS-5-43-1002 → 4460-5-43-1002 (14 chars)
    2. Budget = 18 - len(base)
    3. Try '-L' + zero_padded(item_num) to fill budget
    4. If item_num too large: use '-{first_digit}L{rest}' format
    5. If still too large: shorten base by removing second segment,
       prefix with '0', and retry
    6. Hard cap at 18 chars
    """
    s = spir_no.strip()
    s = re.sub(r'^VEN-', '', s, flags=re.IGNORECASE)
    parts = s.split('-')
    kept = [p.strip() for p in parts if p.strip() and re.match(r'^\d', p.strip())]
    base = '-'.join(kept)  # e.g. '4460-5-43-1002' = 14 chars

    budget = 18 - len(base)
    item_str = str(item_num)

    # Attempt 1: '-L' + zero_padded (needs budget-2 chars for the number)
    num_width = budget - 2  # e.g. 4-2 = 2
    if num_width > 0 and len(item_str) <= num_width:
        result = base + '-L' + item_str.zfill(num_width)
        if len(result) <= 18:
            return result

    # Attempt 2: '-{d1}L{rest}' — first digit before L, rest after
    if len(item_str) >= 2:
        suffix = '-' + item_str[0] + 'L' + item_str[1:]
        result = base + suffix
        if len(result) <= 18:
            return result

    # Attempt 3: Shorten base — remove second hyphen-segment, add leading '0'
    base_parts = base.split('-')
    if len(base_parts) >= 4:
        # e.g. ['4391','5','43','0006'] → '04391-43-0006' (12 chars)
        short_base = '0' + base_parts[0] + '-' + '-'.join(base_parts[2:])
        budget3 = 18 - len(short_base)
        item_str3 = item_str
        if len(item_str3) >= 2:
            suffix3 = '-' + item_str3[0] + 'L' + item_str3[1:]
        else:
            suffix3 = '-L' + item_str3.zfill(budget3 - 2)
        result = short_base + suffix3
        if len(result) <= 18:
            return result

    # Fallback: hard truncate
    return (base + '-L' + item_str)[:18]


def build_position_number(item_num: int) -> str:
    """
    POSITION NUMBER = item_num * 10, zero-padded to 4 digits.
    item 1 → '0010', item 2 → '0020', item 10 → '0100'
    """
    return str(item_num * 10).zfill(4)


# ── Metadata reader ───────────────────────────────────────────────────────────

def _read_metadata(ws) -> dict:
    """
    Read equipment metadata from header rows (1-7).
    EQPT MAKE = MANUFACTURER (the equipment maker, from row 3 right-side)
    """
    meta = {
        "equipment_desc": "",
        "manufacturer":   "",   # → EQPT MAKE
        "supplier":       "",
        "project_name":   "",
        "spir_type":      _detect_spir_type(ws),
    }

    # Primary positions
    v = _v(ws, 2, 24)
    if v: meta["project_name"] = v
    v = _v(ws, 3, 25)
    if v: meta["manufacturer"] = v
    v = _v(ws, 4, 23)
    if v and not meta["supplier"]: meta["supplier"] = v

    # Scan for labels if primary positions empty
    for ri in range(2, 5):
        for ci in range(1, min(ws.max_column + 1, 30)):
            label = _v(ws, ri, ci).lower()
            if "manufacturer:" in label or label == "manufacturer:":
                for dc in range(1, 8):
                    v = _v(ws, ri, ci + dc)
                    if v and not any(kw in v.lower() for kw in
                                     ["manufacturer", "supplier", "equipment", "authority"]):
                        meta["manufacturer"] = meta["manufacturer"] or v
                        break
            if "supplier:" in label or label == "supplier:":
                for dc in range(1, 8):
                    v = _v(ws, ri, ci + dc)
                    if v and not any(kw in v.lower() for kw in
                                     ["supplier", "manufacturer", "equipment"]):
                        meta["supplier"] = meta["supplier"] or v
                        break
            if "equipment:" in label or label == "equipment:":
                for dc in range(1, 8):
                    v = _v(ws, ri, ci + dc)
                    if v and not any(kw in v.lower() for kw in ["equipment"]):
                        meta["project_name"] = meta["project_name"] or v
                        break

    return meta


# ── Column map builder ────────────────────────────────────────────────────────

def _build_col_map(ws, item_col: int) -> dict[int, str]:
    """Build data column → field name mapping from row 7 labels and row 6 headers."""
    col_map: dict[int, str] = {item_col: "item_num"}

    for c in range(item_col + 1, ws.max_column + 1):
        label7 = _v(ws, 7, c).lower().replace(" ", "")
        if label7 in _COL_LABEL_TO_FIELD:
            col_map[c] = _COL_LABEL_TO_FIELD[label7]
            continue
        label6 = _v(ws, 6, c).lower()
        for kw, field in _HEADER_FALLBACK.items():
            if kw in label6:
                col_map[c] = field
                break

    # Ensure unit_price is always mapped (col 22 in standard layout)
    if not any(f == "unit_price" for f in col_map.values()):
        # Find col 22 or the col after CURRENCY
        currency_col = None
        for c, f in col_map.items():
            if f == "currency":
                currency_col = c
                break
        if currency_col:
            col_map[currency_col + 1] = "unit_price"

    if not any(f == "qty_identical" for f in col_map.values()):
        col_map[item_col + 1] = "qty_identical"
    if not any(f == "desc" for f in col_map.values()):
        col_map[item_col + 2] = "desc"

    return col_map


# ── Item data reader ──────────────────────────────────────────────────────────

def _read_item_data(ws, r: int, col_map: dict[int, str],
                    sheet_name: str, spir_no: str, spir_type: str) -> dict:
    """Read one spare row into a dict using the column map."""
    data: dict = {
        "sheet":     sheet_name.strip().upper(),
        "spir_no":   spir_no,
        "spir_type": _normalise_spir_type(spir_type),
    }
    for c, field in col_map.items():
        if field == "item_num": continue
        val = _raw(ws, r, c)
        if val is None: continue
        s = str(val).strip()
        if not s or s == ".": continue
        if field == "qty_identical":
            try: val = int(float(s))
            except: pass
        elif field == "unit_price":
            try: val = float(s)
            except: pass
        data[field] = val

    SKIP = {"na", "tba", "n/a", "tbd", "-", "--", "none", "nil"}

    def _clean(v) -> str:
        if v is None: return ""
        s = str(v).strip()
        return "" if s.lower() in SKIP else s

    desc        = _clean(data.get("desc", ""))
    mfr_part    = _clean(data.get("mfr_part_no", ""))
    supp_part   = _clean(data.get("supplier_part_no", ""))
    supplier_nm = _clean(data.get("supplier_name", ""))

    # MANUFACTURER PART NUMBER: col 11 (10A) when real, else col 12 (10B)
    if mfr_part:
        data["mfr_part_no"] = mfr_part
    elif supp_part:
        data["mfr_part_no"] = supp_part
    else:
        data.pop("mfr_part_no", None)

    final_mfr = data.get("mfr_part_no", "")

    # NEW DESCRIPTION = desc + mfr_part_no + supplier_name
    if desc:
        parts = [desc]
        if final_mfr:   parts.append(str(final_mfr).strip())
        if supplier_nm: parts.append(supplier_nm)
        data["new_desc"] = ", ".join(parts)

    # QAR price
    unit_price = data.get("unit_price")
    currency   = data.get("currency", "")
    if unit_price is not None and currency:
        try:
            from services.currency_service import to_qar
            qar = to_qar(float(unit_price), currency)
            if qar is not None:
                data["unit_price_qar"] = qar
        except Exception:
            pass

    return data


# ── Sheet classifier ──────────────────────────────────────────────────────────

def _classify_sheet(sn: str, ws) -> str:
    lower = sn.lower()
    if "validation" in lower: return "skip"

    has_item_header = any(
        "item number" in _v(ws, 6, c).lower()
        for c in range(1, ws.max_column + 1)
    )
    has_desc = any(
        "description" in _v(ws, r, c).lower()
        for r in [6, 7] for c in range(1, ws.max_column + 1)
    )

    if "continuation" not in lower and "annexure" not in lower:
        return "main"
    if "annexure" in lower:
        return "main"

    # Detect truly empty continuation sheets (blank templates).
    # These have sequential numbers in col 3 (row labels) but no actual
    # tag headers (cols 4+) and no flag data in data rows.
    # Such sheets must be skipped entirely to avoid spurious item linkage.
    if "continuation" in lower:
        has_tags = any(
            _is_tag(_v(ws, 1, c))
            for c in range(4, min(ws.max_column + 1, 50))
        )
        if not has_tags:
            # Check for any real flag data in data rows (cols 4-40, excluding col 41 remarks)
            has_real_data = False
            for r in range(8, min(ws.max_row + 1, 20)):
                for c in range(4, min(ws.max_column + 1, 41)):
                    v = ws.cell(r, c).value
                    if v is not None and str(v).strip() not in ('', '0', '.', 'None'):
                        try:
                            if float(str(v)) > 0:
                                has_real_data = True
                                break
                        except (ValueError, TypeError):
                            has_real_data = True
                            break
                if has_real_data:
                    break
            if not has_real_data:
                return "skip"   # blank template — ignore completely

    if has_desc:
        return "cont_full"
    if has_item_header:
        return "cont_marker"

    has_tags = any(
        _is_tag(_v(ws, 1, c))
        for c in range(3, min(ws.max_column + 1, 50))
    )
    return "cont_overflow_tagged" if has_tags else "cont_overflow_notag"


# ── Row builder ───────────────────────────────────────────────────────────────

def _build_output_row(row_data: dict, is_header: bool, spir_no: str,
                      item_num: Optional[int] = None) -> dict:
    """
    Build a standardised output row dict with all 27 columns.
    is_header=True  → summary row (EQPT_QTY filled, ITEM NUMBER empty)
    is_header=False → detail row (ITEM NUMBER filled, EQPT_QTY empty)
    """
    out = {col: None for col in OUTPUT_COLS}

    out['SPIR NO']    = row_data.get('spir_no', spir_no)
    out['TAG NO']     = row_data.get('tag_no')
    out['EQPT MAKE']  = row_data.get('manufacturer', '')   # ← manufacturer = EQPT MAKE
    out['EQPT MODEL'] = row_data.get('model', '')
    out['EQPT SR NO'] = row_data.get('serial', '')
    out['SPIR ERROR'] = 0
    out['SHEET']      = row_data.get('sheet', '')
    out['SPIR TYPE']  = _normalise_spir_type(row_data.get('spir_type', ''))

    if is_header:
        # Summary row: equipment-level info
        out['EQPT QTY']           = row_data.get('eqpt_qty', 1)
        out['DESCRIPTION OF PARTS'] = row_data.get('desc', '')
        # POSITION NUMBER on header = '0010' (default first position)
        out['POSITION NUMBER']    = '0010'
    else:
        # Detail row: spare part info
        if item_num is not None:
            out['ITEM NUMBER']                   = item_num
            out['POSITION NUMBER']               = build_position_number(item_num)
            out['OLD MATERIAL NUMBER/SPF NUMBER'] = build_old_material_number(
                spir_no, item_num)
        out['QUANTITY IDENTICAL PARTS FITTED'] = row_data.get('qty_identical', 1)
        out['DESCRIPTION OF PARTS']            = row_data.get('desc', '')
        out['NEW DESCRIPTION OF PARTS']        = row_data.get('new_desc', '')
        out['DWG NO INCL POSN NO']             = row_data.get('dwg_no', '')
        out['MANUFACTURER PART NUMBER']        = row_data.get('mfr_part_no', '')
        out['MATERIAL SPECIFICATION']          = row_data.get('material_spec', '')
        out['SUPPLIER OCM NAME']               = row_data.get('supplier_name', '')
        out['CURRENCY']                        = row_data.get('currency', '')
        out['UNIT PRICE']                      = row_data.get('unit_price')
        out['UNIT PRICE (QAR)']                = row_data.get('unit_price_qar')
        out['DELIVERY TIME IN WEEKS']          = row_data.get('delivery')
        out['MIN MAX STOCK LVLS QTY']          = row_data.get('min_max')
        out['UNIT OF MEASURE']                 = row_data.get('uom', '')
        out['SAP NUMBER']                      = row_data.get('sap_no', '')
        out['CLASSIFICATION OF PARTS']         = row_data.get('classification', '')

    return out


# ── Main sheet extractor ──────────────────────────────────────────────────────

def _extract_main(ws, sheet_name: str, spir_no: str):
    if ws.max_row < 7: return {}, {}, {}

    item_col = None
    for c in range(1, ws.max_column + 1):
        if "item number" in _v(ws, 6, c).lower():
            item_col = c; break

    if item_col is None:
        desc_col = None
        for c in range(1, ws.max_column + 1):
            if "description" in _v(ws, 7, c).lower():
                desc_col = c; break
        if desc_col is not None:
            for candidate in [desc_col - 1, desc_col - 2]:
                if candidate < 1: continue
                vals = [ws.cell(r, candidate).value for r in range(8, min(ws.max_row+1, 16))]
                nums = [int(float(str(v))) for v in vals if v is not None
                        and str(v).strip() not in ('', '.')]
                if len(nums) >= 3 and nums == list(range(nums[0], nums[0]+len(nums))):
                    item_col = candidate; break

    if item_col is None: return {}, {}, {}

    tags_raw, tag_cols = [], []
    for c in range(3, item_col):
        val = _v(ws, 1, c)
        if _is_tag(val):
            tags_raw.append(val)
            tag_cols.append(c)
    if not tags_raw: return {}, {}, {}

    models    = [_v(ws, 4, c) for c in tag_cols]
    serials   = [_v(ws, 6, c) for c in tag_cols]
    eqpt_qtys = []
    for c in tag_cols:
        v = _raw(ws, 7, c)
        try: eqpt_qtys.append(int(float(str(v))) if v is not None else 1)
        except: eqpt_qtys.append(1)

    col_map   = _build_col_map(ws, item_col)
    spir_type = _detect_spir_type(ws)
    meta      = _read_metadata(ws)

    spir_type_cols: dict[int, str] = {}
    for c in range(item_col, min(ws.max_column + 1, item_col + 20)):
        for r in [5, 6]:
            v = _v(ws, r, c).lower()
            if "initial spare" in v:       spir_type_cols[c] = "Initial Spare Parts"
            elif "normal" in v and "spare" in v: spir_type_cols[c] = "Normal Operating Spares"
            elif "commissioning" in v:     spir_type_cols[c] = "Commissioning Spare Parts"
            elif "life cycle" in v:        spir_type_cols[c] = "Life Cycle Spare Parts"

    tag_meta: dict[str, dict] = {}
    for i, tag_raw in enumerate(tags_raw):
        for tag in _expand_tags(tag_raw):
            if not tag: continue
            tag_meta[tag] = {
                "model":        models[i]    if i < len(models)    else "",
                "serial":       serials[i]   if i < len(serials)   else "",
                "eqpt_qty":     eqpt_qtys[i] if i < len(eqpt_qtys) else 1,
                "sheet":        sheet_name.strip().upper(),
                "manufacturer": meta["manufacturer"],
                "project_name": meta["project_name"],
            }

    items:     dict[tuple, dict]      = {}
    tag_items: dict[str, list[tuple]] = {}

    for r in range(8, ws.max_row + 1):
        item_val = _raw(ws, r, item_col)
        if item_val is None: continue
        try: item_num = int(float(str(item_val)))
        except: continue
        if item_num <= 0: continue

        key = (sheet_name, item_num)

        row_spir_type = spir_type
        # Check each SPIR type column — only override if cell value is genuinely > 0
        for tc, type_label in spir_type_cols.items():
            cell_val = _raw(ws, r, tc)
            if cell_val is not None and cell_val not in (0, "0", "", "."):
                try:
                    if float(str(cell_val)) > 0:
                        row_spir_type = type_label
                        break
                except (ValueError, TypeError):
                    pass

        data = _read_item_data(ws, r, col_map, sheet_name, spir_no, row_spir_type)
        if data.get("desc") or data.get("mfr_part_no"):
            items[key] = data

        for i, (tag_raw, tc) in enumerate(zip(tags_raw, tag_cols)):
            cell_val = _raw(ws, r, tc)
            if cell_val is None: continue
            try:
                if int(float(str(cell_val))) <= 0: continue
            except:
                s = str(cell_val).strip()
                if not s or s in ("0", "."): continue
            for tag in _expand_tags(tag_raw):
                if not tag: continue
                tag_items.setdefault(tag, []).append(key)

    return items, tag_items, tag_meta


# ── Continuation sheet extractor ──────────────────────────────────────────────

def _extract_cont_marker(ws, sheet_name: str, main_sheets: list[str],
                          master_items: dict) -> dict:
    tags_raw, tag_cols = [], []
    tag_models:     dict[int, str] = {}
    tag_serials:    dict[int, str] = {}
    tag_eqpt_qtys:  dict[int, int] = {}

    for c in range(4, ws.max_column + 1):
        val = _v(ws, 1, c)
        if _is_tag(val):
            tags_raw.append(val)
            tag_cols.append(c)
            tag_models[c]  = _v(ws, 4, c)
            tag_serials[c] = _v(ws, 6, c)
            try:
                eq = int(float(str(_raw(ws, 7, c) or 1)))
                tag_eqpt_qtys[c] = eq if eq > 0 else 1
            except:
                tag_eqpt_qtys[c] = 1

    if not tags_raw: return {}

    cont_sheet_upper = sheet_name.strip().upper()
    tag_items:    dict[str, list[tuple]] = {}
    tag_meta_out: dict[str, dict]        = {}

    for r in range(8, ws.max_row + 1):
        item_val = _raw(ws, r, 3)
        if item_val is None: continue
        try: item_num = int(float(str(item_val)))
        except: continue
        if item_num <= 0: continue

        item_exists = any((ms, item_num) in master_items for ms in main_sheets)
        if not item_exists: continue

        for tag_raw, tc in zip(tags_raw, tag_cols):
            cell_val = _raw(ws, r, tc)
            if cell_val is None: continue
            try:
                if int(float(str(cell_val))) <= 0: continue
            except:
                s = str(cell_val).strip()
                if not s or s in ("0", "."): continue

            for tag in _expand_tags(tag_raw):
                if not tag: continue
                key = (cont_sheet_upper, item_num, tc)
                tag_items.setdefault(tag, []).append(key)
                meta_key = (tag, tc)
                if meta_key not in tag_meta_out:
                    tag_meta_out[meta_key] = {
                        "model":        tag_models.get(tc, ""),
                        "serial":       tag_serials.get(tc, ""),
                        "eqpt_qty":     tag_eqpt_qtys.get(tc, 1),
                        "sheet":        cont_sheet_upper,
                        "manufacturer": "",
                        "project_name": "",
                    }

    return {"tag_items": tag_items, "tag_meta": tag_meta_out}


def _extract_cont_overflow_tagged(ws, sheet_name: str, main_sheets: list[str],
                                   master_items: dict) -> tuple[dict, dict]:
    tags_raw: list[str] = []
    for c in range(3, min(ws.max_column + 1, 15)):
        val = _v(ws, 1, c)
        if _is_tag(val):
            tags_raw.append(val)

    if not tags_raw: return {}, {}

    item_nums: list[int] = []
    for r in range(8, ws.max_row + 1):
        v = _raw(ws, r, 4)
        if v is None: continue
        try:
            n = int(float(str(v)))
            if n > 0: item_nums.append(n)
        except: pass

    tag_items: dict[str, list[tuple]] = {}
    tag_meta:  dict[str, dict]        = {}
    for tag_raw in tags_raw:
        for tag in _expand_tags(tag_raw):
            if not tag: continue
            tag_meta[tag] = {"model": "", "serial": "", "eqpt_qty": 1,
                             "sheet": sheet_name.strip().upper(),
                             "manufacturer": "", "project_name": ""}
            for item_num in item_nums:
                for msheet in main_sheets:
                    key = (msheet, item_num)
                    if key in master_items:
                        tag_items.setdefault(tag, []).append(key)
                        break
    return tag_items, tag_meta


def _extract_cont_overflow_notag(ws, main_sheets: list[str],
                                  master_items: dict,
                                  existing_tags: dict) -> dict[str, list[tuple]]:
    """
    Handles continuation sheets that have no explicit tag headers but list
    item numbers in col 3. Only adds items that GENUINELY exist in master_items
    AND are NOT already linked to the tag. Empty/template continuation sheets
    (whose col-3 numbers are just row-index labels with no real item data)
    are automatically skipped because no matches will be found.
    """
    item_nums: list[int] = []
    for r in range(8, ws.max_row + 1):
        v = _raw(ws, r, 3)
        if v is None: continue
        try:
            n = int(float(str(v)))
            if n > 0: item_nums.append(n)
        except: pass

    # Guard: if NONE of the col-3 values match any master_item, this sheet
    # is a blank template (row-index labels only) — skip entirely
    any_real_item = any(
        (ms, n) in master_items
        for n in item_nums
        for ms in main_sheets
    )
    if not any_real_item:
        return {}

    additions: dict[str, list[tuple]] = {}
    for tag in existing_tags:
        existing_keys = set(existing_tags.get(tag, []))
        for item_num in item_nums:
            for msheet in main_sheets:
                key = (msheet, item_num)
                if key in master_items and key not in existing_keys:
                    additions.setdefault(tag, []).append(key)
                    break
    return additions


# ── Annexure-style extractor ──────────────────────────────────────────────────

def _is_annexure_style(wb) -> bool:
    for sn in wb.sheetnames:
        if "validation" in sn.lower() or "continuation" in sn.lower(): continue
        ws = wb[sn]
        for c in range(1, min(ws.max_column + 1, 12)):
            if "annexure" in _v(ws, 1, c).lower():
                return True
    return False


def _extract_annexure_style(wb, spir_no: str) -> dict:
    """
    Annexure format: single MAIN SHEET + Annexure sheets.
    Tags live in Annexure sheets (rows), not as column headers.
    Output follows same 27-column structure as other formats.
    """
    main_ws, main_name = None, ""
    for sn in wb.sheetnames:
        if "validation" in sn.lower() or "annexure" in sn.lower(): continue
        main_ws = wb[sn]; main_name = sn; break
    if main_ws is None:
        return _empty_result(spir_no)

    # Detect item col
    item_col = None
    for c in range(1, main_ws.max_column + 1):
        if "item number" in _v(main_ws, 6, c).lower():
            item_col = c; break
    if item_col is None: item_col = 9

    # Annexure column flags in main sheet
    annexure_cols: list[tuple[int, str]] = []
    for c in range(1, item_col):
        v = _v(main_ws, 1, c)
        if "annexure" in v.lower():
            annexure_cols.append((c, v.strip()))

    col_map   = _build_col_map(main_ws, item_col)
    spir_type = _detect_spir_type(main_ws)
    meta      = _read_metadata(main_ws)
    manufacturer = meta["manufacturer"]

    # Read items from main sheet
    items: dict[int, dict] = {}
    item_to_annexures: dict[int, list[str]] = {}

    for r in range(8, main_ws.max_row + 1):
        item_val = _raw(main_ws, r, item_col)
        if item_val is None: continue
        try: item_num = int(float(str(item_val)))
        except: continue
        if item_num <= 0: continue

        data = _read_item_data(main_ws, r, col_map, main_name, spir_no, spir_type)
        if not data.get("desc"): continue
        # Fill manufacturer from sheet meta if not in item
        if not data.get("manufacturer"):
            data["manufacturer"] = manufacturer
        items[item_num] = data
        item_to_annexures[item_num] = []
        for ac, ann_name in annexure_cols:
            cv_val = _raw(main_ws, r, ac)
            if cv_val is not None and str(cv_val).strip() not in ("", "0", "."):
                item_to_annexures[item_num].append(ann_name)

    # Read tags from annexure sheets
    annexure_tags: dict[str, list[dict]] = {}
    for sn in wb.sheetnames:
        if "annexure" not in sn.lower(): continue
        ws = wb[sn]
        tags = []
        # Find header row containing 'Sr. No' or 'Sr.No'
        hdr_row = None
        for ri in range(1, min(ws.max_row + 1, 10)):
            s = '|'.join(str(ws.cell(ri, ci).value or '') for ci in range(1, 10)).lower()
            if 'sr. no' in s or 'sr.no' in s:
                hdr_row = ri; break
        if hdr_row:
            for ri in range(hdr_row + 2, ws.max_row + 1):
                tag = _v(ws, ri, 5)   # col 5 = valve tag
                serial = _v(ws, ri, 23)  # col 23 = mfr serial
                model  = _v(ws, ri, 24)  # col 24 = mfr model
                if tag and _is_tag(tag):
                    tags.append({
                        "tag":    tag,
                        "serial": serial,
                        "model":  model,
                    })
        else:
            # Fallback: scan for tags in col 5
            for r in range(5, ws.max_row + 1):
                tag = _v(ws, r, 5)
                if _is_tag(tag):
                    tags.append({
                        "tag":    tag,
                        "serial": _v(ws, r, 23),
                        "model":  _v(ws, r, 24),
                    })
        if tags: annexure_tags[sn] = tags

    # Build output rows (27-column format)
    output_rows: list[dict] = []
    spare_count = 0
    all_tags: set[str] = set()

    for ac, ann_name in annexure_cols:
        ann_ws_name = next((k for k in annexure_tags
                            if ann_name.lower() in k.lower()), None)
        if not ann_ws_name:
            continue
        ann_items = [n for n, anns in item_to_annexures.items()
                     if ann_name in anns]

        # EQPT QTY for this annexure = unit count from main sheet row 7
        ann_col_idx = ac
        eqpt_qty_val = _raw(main_ws, 7, ann_col_idx)
        try:
            eqpt_qty = int(float(str(eqpt_qty_val))) if eqpt_qty_val else len(annexure_tags[ann_ws_name])
        except:
            eqpt_qty = len(annexure_tags[ann_ws_name])

        # ── Global rows: one header + one detail per item, NO tag ────────────
        # These appear once per annexure group, before all the per-tag rows.
        # Expected structure: global_header → global_detail → (per-tag header → per-tag detail) × N
        for item_num in ann_items:
            d = items.get(item_num, {})
            if not d:
                continue

            # Global summary row (no TAG NO, no EQPT MODEL, no EQPT SR NO)
            global_hdr = {col: None for col in OUTPUT_COLS}
            global_hdr['SPIR NO']              = spir_no
            global_hdr['EQPT MAKE']            = manufacturer
            global_hdr['EQPT QTY']             = eqpt_qty
            global_hdr['POSITION NUMBER']      = build_position_number(item_num)
            global_hdr['DESCRIPTION OF PARTS'] = meta.get('project_name', '')
            global_hdr['SPIR ERROR']           = 0
            global_hdr['SHEET']                = main_name.strip().upper()
            global_hdr['SPIR TYPE']            = _normalise_spir_type(spir_type)
            output_rows.append(global_hdr)

            # Global detail row (no TAG NO)
            global_det = {col: None for col in OUTPUT_COLS}
            global_det['SPIR NO']                         = spir_no
            global_det['EQPT MAKE']                       = manufacturer
            global_det['QUANTITY IDENTICAL PARTS FITTED'] = d.get('qty_identical', 1)
            global_det['ITEM NUMBER']                     = item_num
            global_det['POSITION NUMBER']                 = build_position_number(item_num)
            global_det['OLD MATERIAL NUMBER/SPF NUMBER']  = build_old_material_number(spir_no, item_num)
            global_det['DESCRIPTION OF PARTS']            = d.get('desc', '')
            global_det['NEW DESCRIPTION OF PARTS']        = d.get('new_desc', '')
            global_det['DWG NO INCL POSN NO']             = d.get('dwg_no', '')
            global_det['MANUFACTURER PART NUMBER']        = d.get('mfr_part_no', '')
            global_det['MATERIAL SPECIFICATION']          = d.get('material_spec', '')
            global_det['SUPPLIER OCM NAME']               = d.get('supplier_name', '')
            global_det['CURRENCY']                        = d.get('currency', '')
            global_det['UNIT PRICE']                      = d.get('unit_price')
            global_det['UNIT PRICE (QAR)']                = d.get('unit_price_qar')
            global_det['DELIVERY TIME IN WEEKS']          = d.get('delivery')
            global_det['MIN MAX STOCK LVLS QTY']          = d.get('min_max')
            global_det['UNIT OF MEASURE']                 = d.get('uom', '')
            global_det['SAP NUMBER']                      = d.get('sap_no', '')
            global_det['CLASSIFICATION OF PARTS']         = d.get('classification', '')
            global_det['SPIR ERROR']                      = 0
            global_det['SHEET']                           = d.get('sheet', main_name).strip().upper()
            global_det['SPIR TYPE']                       = _normalise_spir_type(
                d.get('spir_type', spir_type))
            output_rows.append(global_det)

        # ── Per-tag rows: header + details for each tag in this annexure ─────
        for ti in annexure_tags[ann_ws_name]:
            tag = ti["tag"]
            all_tags.add(tag)

            # ── Per-tag header (summary) row ────────────────────────────────
            header_row = {col: None for col in OUTPUT_COLS}
            header_row['SPIR NO']             = spir_no
            header_row['TAG NO']              = tag
            header_row['EQPT MAKE']           = manufacturer
            header_row['EQPT MODEL']          = ti.get('model', '')
            header_row['EQPT SR NO']          = ti.get('serial', '')
            header_row['EQPT QTY']            = eqpt_qty
            header_row['POSITION NUMBER']     = '0010'
            header_row['DESCRIPTION OF PARTS'] = meta.get('project_name', '')
            header_row['SPIR ERROR']          = 0
            header_row['SHEET']               = main_name.strip().upper()
            header_row['SPIR TYPE']           = _normalise_spir_type(spir_type)
            output_rows.append(header_row)

            # ── Per-tag detail rows ───────────────────────────────────────────
            for item_num in ann_items:
                d = items.get(item_num, {})
                if not d: continue

                detail_row = {col: None for col in OUTPUT_COLS}
                detail_row['SPIR NO']                          = spir_no
                detail_row['TAG NO']                           = tag
                detail_row['EQPT MAKE']                        = manufacturer
                detail_row['EQPT MODEL']                       = ti.get('model', '')
                detail_row['EQPT SR NO']                       = ti.get('serial', '')
                detail_row['QUANTITY IDENTICAL PARTS FITTED']  = d.get('qty_identical', 1)
                detail_row['ITEM NUMBER']                      = item_num
                detail_row['POSITION NUMBER']                  = build_position_number(item_num)
                detail_row['OLD MATERIAL NUMBER/SPF NUMBER']   = build_old_material_number(spir_no, item_num)
                detail_row['DESCRIPTION OF PARTS']             = d.get('desc', '')
                detail_row['NEW DESCRIPTION OF PARTS']         = d.get('new_desc', '')
                detail_row['DWG NO INCL POSN NO']              = d.get('dwg_no', '')
                detail_row['MANUFACTURER PART NUMBER']         = d.get('mfr_part_no', '')
                detail_row['MATERIAL SPECIFICATION']           = d.get('material_spec', '')
                detail_row['SUPPLIER OCM NAME']                = d.get('supplier_name', '')
                detail_row['CURRENCY']                         = d.get('currency', '')
                detail_row['UNIT PRICE']                       = d.get('unit_price')
                detail_row['UNIT PRICE (QAR)']                 = d.get('unit_price_qar')
                detail_row['DELIVERY TIME IN WEEKS']           = d.get('delivery')
                detail_row['MIN MAX STOCK LVLS QTY']           = d.get('min_max')
                detail_row['UNIT OF MEASURE']                  = d.get('uom', '')
                detail_row['SAP NUMBER']                       = d.get('sap_no', '')
                detail_row['CLASSIFICATION OF PARTS']          = d.get('classification', '')
                detail_row['SPIR ERROR']                       = 0
                detail_row['SHEET']                            = d.get('sheet', main_name).strip().upper()
                detail_row['SPIR TYPE']                        = _normalise_spir_type(
                    d.get('spir_type', spir_type))
                output_rows.append(detail_row)
                spare_count += 1

    return {
        "format":          "SPIR_ANNEXURE",
        "spir_no":         spir_no,
        "equipment":       meta["project_name"],
        "manufacturer":    manufacturer,
        "supplier":        meta.get("supplier", ""),
        "spir_type":       _normalise_spir_type(spir_type),
        "eqpt_qty":        len(all_tags),
        "spare_items":     spare_count,
        "total_tags":      len(all_tags),
        "annexure_count":  len(annexure_tags),
        "annexure_stats":  {k: len(v) for k, v in annexure_tags.items()},
        "rows":            output_rows,
        "output_cols":     OUTPUT_COLS,
    }


# ── Empty result ──────────────────────────────────────────────────────────────

def _empty_result(spir_no: str) -> dict:
    return {
        "format": "UNKNOWN", "rows": [], "total_tags": 0, "spare_items": 0,
        "spir_no": spir_no, "eqpt_qty": 0, "annexure_count": 0,
        "annexure_stats": {}, "manufacturer": "", "supplier": "",
        "spir_type": "", "equipment": "", "output_cols": OUTPUT_COLS,
    }


# ── Main entry point ──────────────────────────────────────────────────────────

def extract_workbook(wb, spir_no: str = "", spir_filename: str = "") -> dict:
    """
    Main entry point. Auto-detects format and extracts all SPIR data.
    Returns dict with 'rows' (list of 27-column dicts) and metadata.
    """
    if not spir_no:
        spir_no = _detect_spir_no(wb)
    if not spir_no and spir_filename:
        m = re.search(
            r'(VEN-[A-Z0-9]+-[A-Z0-9]+-\d+-\d+-[A-Z0-9.\-]+|'
            r'\d{4}-[A-Z]{2}-\d{2}-\d{2}-\d{2}-\d{3}[^\s_]*)',
            spir_filename, re.IGNORECASE)
        if m: spir_no = m.group(1)

    if _is_annexure_style(wb):
        log.info("SPIR format: ANNEXURE-STYLE")
        return _extract_annexure_style(wb, spir_no)

    log.info("extract_workbook: spir_no=%r  sheets=%s", spir_no, wb.sheetnames)

    master_items:   dict[tuple, dict]       = {}
    all_tag_items:  dict[str, list[tuple]]  = {}
    all_tag_meta:   dict[str, dict]         = {}
    main_sheets:    list[str]               = []
    annexure_stats: dict[str, int]          = {}
    manufacturer = supplier = spir_type = equipment = ""

    # Pass 1: main / full sheets
    for sheet_name in wb.sheetnames:
        ws       = wb[sheet_name]
        category = _classify_sheet(sheet_name, ws)
        if category in ("skip", "cont_marker",
                        "cont_overflow_tagged", "cont_overflow_notag"):
            continue

        items, tag_items, tag_meta = _extract_main(ws, sheet_name, spir_no)
        if not items: continue

        master_items.update(items)
        main_sheets.append(sheet_name)

        for tag, keys in tag_items.items():
            all_tag_items.setdefault(tag, []).extend(keys)
            if tag not in all_tag_meta:
                all_tag_meta[tag] = tag_meta.get(tag, {})

        n = sum(len(v) for v in tag_items.values())
        if n: annexure_stats[sheet_name] = n

        if not manufacturer:
            meta = _read_metadata(ws)
            manufacturer = meta.get("manufacturer", "")
            equipment    = meta.get("project_name", "")
        if not spir_type:
            for d in items.values():
                t = d.get("spir_type", "")
                if t: spir_type = t; break

    # Pass 2: continuation sheets
    for sheet_name in wb.sheetnames:
        ws       = wb[sheet_name]
        category = _classify_sheet(sheet_name, ws)

        if category == "cont_marker":
            cont_result = _extract_cont_marker(ws, sheet_name, main_sheets, master_items)
            tag_items_c = cont_result.get("tag_items", {})
            tag_meta_c  = cont_result.get("tag_meta", {})

            for tag, keys in tag_items_c.items():
                all_tag_items.setdefault(tag, []).extend(keys)
                for k in keys:
                    col = k[2] if len(k) > 2 else None
                    mk  = (tag, col)
                    if mk not in all_tag_meta:
                        me = dict(tag_meta_c.get(mk, tag_meta_c.get(tag, {})))
                        me["manufacturer"] = me.get("manufacturer") or manufacturer
                        all_tag_meta[mk] = me
                if tag not in all_tag_meta:
                    me = dict(tag_meta_c.get(tag, {}))
                    me["manufacturer"] = me.get("manufacturer") or manufacturer
                    all_tag_meta[tag] = me

            n = sum(len(v) for v in tag_items_c.values())
            if n: annexure_stats[sheet_name] = n

        elif category == "cont_overflow_tagged":
            additions, tag_meta = _extract_cont_overflow_tagged(
                ws, sheet_name, main_sheets, master_items)
            for tag, keys in additions.items():
                all_tag_items.setdefault(tag, []).extend(keys)
                if tag not in all_tag_meta:
                    all_tag_meta[tag] = tag_meta.get(tag, {})
            n = sum(len(v) for v in additions.values())
            if n: annexure_stats[sheet_name] = n

        elif category == "cont_overflow_notag":
            additions = _extract_cont_overflow_notag(
                ws, main_sheets, master_items, all_tag_items)
            for tag, keys in additions.items():
                all_tag_items[tag] = all_tag_items.get(tag, []) + keys

    # Build output rows
    from collections import defaultdict
    tag_sheet_items: dict[tuple, list[tuple]] = defaultdict(list)
    tag_sheet_order: list[tuple] = []

    for tag, keys in all_tag_items.items():
        seen: set = set()
        for key in keys:
            if key not in seen:
                seen.add(key)
                ts = (tag, key[0], key[2]) if len(key) > 2 else (tag, key[0])
                tag_sheet_items[ts].append(key)
                if ts not in tag_sheet_order:
                    tag_sheet_order.append(ts)

    output_rows:  list[dict] = []
    spare_count = 0

    def _get_item_data(key: tuple) -> Optional[dict]:
        if key in master_items:
            return master_items[key]
        item_num = key[1]
        for ms in main_sheets:
            alt = (ms, item_num)
            if alt in master_items:
                return master_items[alt]
        return None

    for ts_key in tag_sheet_order:
        tag   = ts_key[0]
        sheet = ts_key[1]
        keys  = tag_sheet_items[ts_key]
        if not keys: continue

        ts_col = ts_key[2] if len(ts_key) > 2 else None
        meta   = all_tag_meta.get((tag, ts_col)) or all_tag_meta.get(tag, {})
        fi     = _get_item_data(keys[0]) or {}

        output_sheet = sheet.strip().upper()
        eqpt_qty     = meta.get("eqpt_qty", 1)

        # ── Header row ────────────────────────────────────────────────────────
        header_row = {col: None for col in OUTPUT_COLS}
        header_row['SPIR NO']              = spir_no
        header_row['TAG NO']               = tag
        header_row['EQPT MAKE']            = meta.get("manufacturer", manufacturer)
        header_row['EQPT MODEL']           = meta.get("model", "")
        header_row['EQPT SR NO']           = meta.get("serial", "")
        header_row['EQPT QTY']             = eqpt_qty
        header_row['POSITION NUMBER']      = '0010'
        header_row['DESCRIPTION OF PARTS'] = equipment or fi.get("desc", "")
        header_row['SPIR ERROR']           = 0
        header_row['SHEET']                = output_sheet
        header_row['SPIR TYPE']            = _normalise_spir_type(
            fi.get("spir_type", spir_type))
        output_rows.append(header_row)

        # ── Detail rows ───────────────────────────────────────────────────────
        for key in keys:
            d = _get_item_data(key)
            if d is None: continue
            item_num = key[1]

            detail_row = {col: None for col in OUTPUT_COLS}
            detail_row['SPIR NO']                         = spir_no
            detail_row['TAG NO']                          = tag
            detail_row['EQPT MAKE']                       = meta.get("manufacturer", manufacturer)
            detail_row['EQPT MODEL']                      = meta.get("model", "")
            detail_row['EQPT SR NO']                      = meta.get("serial", "")
            detail_row['QUANTITY IDENTICAL PARTS FITTED'] = d.get("qty_identical", 1)
            detail_row['ITEM NUMBER']                     = item_num
            detail_row['POSITION NUMBER']                 = build_position_number(item_num)
            detail_row['OLD MATERIAL NUMBER/SPF NUMBER']  = build_old_material_number(
                spir_no, item_num)
            detail_row['DESCRIPTION OF PARTS']            = d.get("desc", "")
            detail_row['NEW DESCRIPTION OF PARTS']        = d.get("new_desc", "")
            detail_row['DWG NO INCL POSN NO']             = d.get("dwg_no", "")
            detail_row['MANUFACTURER PART NUMBER']        = d.get("mfr_part_no", "")
            detail_row['MATERIAL SPECIFICATION']          = d.get("material_spec", "")
            detail_row['SUPPLIER OCM NAME']               = d.get("supplier_name", "")
            detail_row['CURRENCY']                        = d.get("currency", "")
            detail_row['UNIT PRICE']                      = d.get("unit_price")
            detail_row['UNIT PRICE (QAR)']                = d.get("unit_price_qar")
            detail_row['DELIVERY TIME IN WEEKS']          = d.get("delivery")
            detail_row['MIN MAX STOCK LVLS QTY']          = d.get("min_max")
            detail_row['UNIT OF MEASURE']                 = d.get("uom", "")
            detail_row['SAP NUMBER']                      = d.get("sap_no", "")
            detail_row['CLASSIFICATION OF PARTS']         = d.get("classification", "")
            detail_row['SPIR ERROR']                      = 0
            detail_row['SHEET']                           = output_sheet
            detail_row['SPIR TYPE']                       = _normalise_spir_type(
                d.get("spir_type", spir_type))
            output_rows.append(detail_row)
            spare_count += 1

    all_tags = {ts[0] for ts in tag_sheet_items}

    return {
        "format":          "SPIR_MATRIX",
        "spir_no":         spir_no,
        "equipment":       equipment,
        "manufacturer":    manufacturer,
        "supplier":        supplier,
        "spir_type":       _normalise_spir_type(spir_type),
        "eqpt_qty":        len(all_tags),
        "spare_items":     spare_count,
        "total_tags":      len(all_tags),
        "annexure_count":  max(0, len(annexure_stats) - 1),
        "annexure_stats":  annexure_stats,
        "rows":            output_rows,
        "output_cols":     OUTPUT_COLS,
    }