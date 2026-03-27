"""
extraction/spir_extractor.py
─────────────────────────────
Qatar Energy SPIR matrix extractor — all verified formats.

OUTPUT COLUMNS (27 total):
  SPIR NO | TAG NO | EQPT MAKE | EQPT MODEL | EQPT SR NO | EQPT QTY |
  QUANTITY IDENTICAL PARTS FITTED | ITEM NUMBER | POSITION NUMBER |
  OLD MATERIAL NUMBER/SPF NUMBER | DESCRIPTION OF PARTS |
  NEW DESCRIPTION OF PARTS | DWG NO INCL POSN NO |
  MANUFACTURER PART NUMBER | MATERIAL SPECIFICATION |
  SUPPLIER OCM NAME | CURRENCY | UNIT PRICE | UNIT PRICE (QAR) |
  DELIVERY TIME IN WEEKS | MIN MAX STOCK LVLS QTY | UNIT OF MEASURE |
  SAP NUMBER | CLASSIFICATION OF PARTS | SPIR ERROR | SHEET | SPIR TYPE

SPIR TYPE values (exactly 4 allowed):
  1. Normal Operating Spares
  2. Initial Spare Parts
  3. Commissioning Spare Parts
  4. Life Cycle Spare Parts

OLD MATERIAL NUMBER rules (REQUIREMENT 1 — v2):
  - Always exactly 18 characters
  - Base = numeric segments of SPIR_NO (after removing VEN- prefix,
    (REV.x) suffixes, and all non-digit-leading segments)
  - Suffix = '-<sheet_index>l<line_index>'
    e.g. sheet 1, line 1  → '-1l1'
         sheet 2, line 17 → '-2l17'
  - If base + suffix > 18: progressively remove second segment from base
    until it fits, then zero-pad / truncate to exactly 18.
  - Max 24 spare lines per sheet (line_index 1..N only, never fabricated).

POSITION NUMBER:
  - Always 4-digit zero-padded: item_num * 10 → '0010', '0020', etc.

EQPT MAKE:
  - Taken from MANUFACTURER field in the SPIR header (row 3, right-side label)
  - NOT the supplier name — the equipment manufacturer
"""
from __future__ import annotations
import re
import logging
from typing import Optional

import openpyxl

log = logging.getLogger(__name__)

# ── Output column order (27 columns) ─────────────────────────────────────────
OUTPUT_COLS = [
    'SPIR NO',
    'TAG NO',
    'EQPT MAKE',
    'EQPT MODEL',
    'EQPT SR NO',
    'EQPT QTY',
    'QUANTITY IDENTICAL PARTS FITTED',
    'ITEM NUMBER',
    'POSITION NUMBER',
    'OLD MATERIAL NUMBER/SPF NUMBER',
    'DESCRIPTION OF PARTS',
    'NEW DESCRIPTION OF PARTS',
    'DWG NO INCL POSN NO',
    'MANUFACTURER PART NUMBER',
    'MATERIAL SPECIFICATION',
    'SUPPLIER OCM NAME',
    'CURRENCY',
    'UNIT PRICE',
    'UNIT PRICE (QAR)',
    'DELIVERY TIME IN WEEKS',
    'MIN MAX STOCK LVLS QTY',
    'UNIT OF MEASURE',
    'SAP NUMBER',
    'CLASSIFICATION OF PARTS',
    'SPIR ERROR',
    'SHEET',
    'SPIR TYPE',
]

# ── SPIR TYPE normalisation ───────────────────────────────────────────────────

def _normalise_spir_type(raw: str) -> str:
    """Normalise any SPIR type string to one of the 4 allowed values."""
    lower = (raw or '').lower().strip()
    if 'commission' in lower:
        return 'Commissioning Spare Parts'
    if 'initial' in lower:
        return 'Initial Spare Parts'
    if 'life' in lower or 'lifecycle' in lower:
        return 'Life Cycle Spare Parts'
    return 'Normal Operating Spares'   # default


# ── Column label → field name (row 7 numeric labels) ─────────────────────────
_COL_LABEL_TO_FIELD = {
    "7":   "qty_identical",
    "8":   "desc",
    "9":   "dwg_no",
    "10a": "mfr_part_no",
    "10b": "supplier_part_no",
    "11a": "material_spec",
    "11b": "material_cert",
    "12a": "supplier_name",
    "12b": "supporting_docs",
    "17":  "currency",
    "18":  "delivery",
    "19":  "min_max",
    "20":  "uom",
    "21":  "sap_no",
    "22":  "classification",
}

_HEADER_FALLBACK = {
    "total no. of identical": "qty_identical",
    "identical parts":        "qty_identical",
    "description of parts":   "desc",
    "dwg no":                 "dwg_no",
    "posn no":                "dwg_no",
    "manufacturer part":      "mfr_part_no",
    "suppliers part":         "supplier_part_no",
    "material spec":          "material_spec",
    "material cert":          "material_cert",
    "supplier/ocm":           "supplier_name",
    "supplier ocm":           "supplier_name",
    "currency":               "currency",
    "unit price":             "unit_price",
    "delivery":               "delivery",
    "stock lvls":             "min_max",
    "unit of measure":        "uom",
    "unit  of measure":       "uom",
    "sap number":             "sap_no",
    "sap no":                 "sap_no",
    "classification":         "classification",
}


def _v(ws, r: int, c: int) -> str:
    val = ws.cell(r, c).value
    if val is None: return ""
    s = str(val).strip()
    return "" if s in (".", "_") else s


def _raw(ws, r: int, c: int):
    return ws.cell(r, c).value


def _is_tag(val: str) -> bool:
    if not val or len(val) > 80: return False
    skip = {"spare", "interchangeability", "record", "spir", "note",
            "authority", "required", "quantities", "authorised",
            "equipment", "tag no", "annexure", "refer", "qatar",
            "equip", "or tag", "or model", "mfr", "serial", "no.",
            "_", "spir number", "na", "n/a", "tba"}
    lower = val.lower().strip()
    if any(lower == w for w in skip): return False
    if any(w in lower for w in ["spare parts", "interchangeability", "record",
                                 "spir number", "authority", "required on site",
                                 "quantities in col"]): return False
    return True


def _expand_tags(raw: str) -> list[str]:
    raw = raw.strip()
    if not raw: return []
    m = re.match(r'^(\d+)\s+to\s+(\d+)$', raw, re.IGNORECASE)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        return [str(i) for i in range(a, b + 1)]
    if "/" in raw or "," in raw:
        parts = [p.strip() for p in re.split(r'[/,]', raw) if p.strip()]
        if len(parts) > 1:
            first = parts[0]
            result = [first]
            for part in parts[1:]:
                if re.match(r'^\d+$', part) and len(part) < len(first):
                    pm = re.match(r'^(.*\D)(\d+)$', first)
                    result.append(pm.group(1) + part if pm else part)
                else:
                    result.append(part)
            return result
    return [raw]


def _expand_serials(raw: str) -> list[str]:
    """
    Expand serial number cell value into a list of individual serials.
    Supports: single, slash-separated, comma-separated, 'X to Y' ranges.
    Returns empty list if raw is empty/None/placeholder.
    """
    if not raw:
        return []
    raw = str(raw).strip()
    if not raw or raw.lower() in ('na', 'n/a', 'tba', 'tbd', '-', '--', 'none', 'nil', '_'):
        return []
    # Range: "X to Y"
    m = re.match(r'^(.+?)\s+to\s+(.+)$', raw, re.IGNORECASE)
    if m:
        try:
            a, b = int(m.group(1).strip()), int(m.group(2).strip())
            return [str(i) for i in range(a, b + 1)]
        except ValueError:
            pass
    # Slash or comma separated
    if '/' in raw or ',' in raw:
        return [p.strip() for p in re.split(r'[/,]', raw) if p.strip()]
    return [raw]


def _detect_spir_type(ws) -> str:
    """
    Detect the SPIR type from the sheet headers.
    Scans ALL rows 1-11 and collects all type mentions.
    If BOTH Initial and Normal are mentioned, prefer Normal Operating Spares
    as sheet-level default (per-item logic overrides with Initial when col 17 > 0).
    """
    found_types: list[str] = []
    for r in range(1, 12):
        for c in range(1, min(ws.max_column + 1, 30)):
            v = _v(ws, r, c).lower()
            if "normal operating" in v or "normal operation" in v:
                found_types.append("Normal Operating Spares")
            elif "initial spare" in v:
                found_types.append("Initial Spare Parts")
            elif "commissioning" in v:
                found_types.append("Commissioning Spare Parts")
            elif "life cycle" in v:
                found_types.append("Life Cycle Spare Parts")
    if not found_types:
        return "Normal Operating Spares"
    if "Normal Operating Spares" in found_types:
        return "Normal Operating Spares"
    return found_types[0]


def _detect_spir_no(wb) -> str:
    pattern = re.compile(
        r'(VEN-\d{4}-[A-Z0-9]+-\d+-\d+-[\w.\-]+|'
        r'\d{4}-[A-Z]{2}-\d{2}-\d{2}-\d{2}-\d{3}[\w.\-]*)',
        re.IGNORECASE)
    for sn in wb.sheetnames:
        if "validation" in sn.lower(): continue
        ws = wb[sn]
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(1, min(50, ws.max_column + 1)):
                v = str(ws.cell(r, c).value or "")
                m = pattern.search(v)
                if m: return m.group(1).strip()
    return ""


# ── Sheet index tracker (for OLD MATERIAL NUMBER sheet_index) ─────────────────
# Maps sheet_name → 1-based sheet index (main sheets only, in encounter order)
_sheet_index_registry: dict[str, int] = {}
_sheet_index_counter: list[int] = [0]   # mutable counter in closure-safe list


def _reset_sheet_registry():
    _sheet_index_registry.clear()
    _sheet_index_counter[0] = 0


def _get_sheet_index(sheet_name: str) -> int:
    """Return the 1-based index for a sheet, registering it on first encounter."""
    key = sheet_name.strip().upper()
    if key not in _sheet_index_registry:
        _sheet_index_counter[0] += 1
        _sheet_index_registry[key] = _sheet_index_counter[0]
    return _sheet_index_registry[key]


# ── OLD MATERIAL NUMBER builder ───────────────────────────────────────────────
# REQUIREMENT 1 (v3): Uppercase L, shared line numbers, max 24, zero-strip rules

def _omn_extract_segments(spir_no: str) -> list[str]:
    """
    Extract numeric-leading segments from a SPIR number.

    Steps:
    1. Remove bracketed suffixes like (REV.2)
    2. Remove VEN- prefix
    3. Split on '-', keep only segments that START with a digit
    4. Drop a trailing SINGLE-digit revision segment if 4+ segments remain
       e.g. 'VEN-4460-KAHS-5-43-1002-2' → ['4460','5','43','1002'] (drop '2')
    """
    s = str(spir_no).strip()
    s = re.sub(r'\s*\(.*?\)\s*', '', s).strip()          # remove (REV.2)
    s = re.sub(r'^VEN-', '', s, flags=re.IGNORECASE)      # remove VEN-
    parts = s.split('-')
    kept = [p.strip() for p in parts
            if p.strip() and re.match(r'^\d', p.strip())]
    # Drop trailing single-digit revision (e.g. the '2' in 4460-KAHS-5-43-1002-2)
    if len(kept) >= 4 and re.match(r'^\d$', kept[-1]):
        kept = kept[:-1]
    return kept


def _omn_strip_leading_zeros(segments: list[str]) -> list[str]:
    """
    Remove leading zeros from numeric segments ONLY when the numeric VALUE
    is unchanged and at least one character is saved.

    Rule: '0610' → '610' ✓ (saves 1 char, value same)
          '0002' → '2'   ✓ (saves 3 chars, value same)
          '053'  → '53'  ✓ (saves 1 char, value same)
          '6100' → keep  ✗ (trailing zero — removing would change value)
          '0'    → keep  ✗ (single zero, can't reduce further)
    """
    result = []
    for seg in segments:
        # Only strip if seg starts with 0 AND has more than 1 digit
        if len(seg) > 1 and seg.startswith('0') and re.match(r'^0+\d+$', seg):
            stripped = seg.lstrip('0') or '0'
            result.append(stripped)      # 0610→610, 0002→2, 053→53
        else:
            result.append(seg)           # 6100→6100 (no leading zero), 0→0
    return result


def build_old_material_number(spir_no: str, sheet_index: int, line_index: int) -> str:
    """
    Build OLD MATERIAL NUMBER — always exactly 18 characters.

    FORMAT:  <numeric_base>-<suffix>  where total = 18 chars

    SUFFIX RULES (uppercase L):
      sheet_index == 1  →  '-L{line:02d}'       e.g. '-L01', '-L28'
      sheet_index  > 1  →  '-{sheet}L{line}'    e.g. '-2L1', '-2L28'

    LINE NUMBER:
      = item_num (the actual item number from the SPIR sheet)
      No hard cap — SPIR sheets may have up to 28 or more items per sheet.
      Compression fires automatically when suffix+base would exceed 18.

    COMPRESSION PRIORITY (when base + suffix > 18):
      1. Original full base (no stripping, no reduction)
      2. Zero-stripped full base (0610→610 before cutting segments)
      3. Original base with progressive interior-segment reduction
      4. Zero-stripped base with progressive interior reduction
      5. Hard-truncate as absolute last resort

    PADDING:
      If result < 18 chars → right-pad with '0'

    LINE NUMBERS ARE SHARED across all tags on the same sheet
    (line_index = item_num, same for all tags referencing the same item)
    """
    line_index = int(line_index)   # no clamping — item_num is the reference

    # Build suffix
    if sheet_index == 1:
        # Sheet 1: '-L{line:02d}' (zero-padded, always 4+ chars)
        suffix = f'-L{line_index:02d}'           # '-L01' .. '-L28'
    else:
        # Sheet 2+: '-{sheet}L{line}'
        suffix = f'-{sheet_index}L{line_index}'  # '-2L1' .. '-2L28'

    base_budget = 18 - len(suffix)

    def _try_fit(segs: list[str], budget: int) -> str | None:
        """Progressive interior-segment reduction. Returns base string or None."""
        candidate = segs[:]
        while candidate:
            base = '-'.join(candidate)
            if len(base) <= budget:
                return base
            # Remove second segment to shorten (preserve first + last group)
            if len(candidate) >= 3:
                candidate = [candidate[0]] + candidate[2:]
            elif len(candidate) == 2:
                candidate = [candidate[0]]
            else:
                break
        return None

    segs_orig     = _omn_extract_segments(spir_no)
    segs_stripped = _omn_strip_leading_zeros(segs_orig)

    # Compression priority (per spec):
    #   1. Original full segments (no stripping, no reduction)
    #   2. Zero-stripped full segments (strip leading zeros before cutting segments)
    #   3. Original segments with progressive interior reduction
    #   4. Stripped segments with progressive interior reduction
    #   5. Hard truncate as last resort
    #
    # PADDING RULE: pad the BASE (not the whole result) so zeros never land
    # after the suffix.  base.ljust(base_budget,'0') + suffix = exactly 18 chars.

    def _assemble(base: str) -> str:
        """Pad base to base_budget then append suffix → always 18 chars."""
        padded = base.ljust(base_budget, '0')[:base_budget]
        return padded + suffix  # len = base_budget + len(suffix) = 18

    # Priority 1: original, full
    base = '-'.join(segs_orig)
    if len(base) <= base_budget:
        return _assemble(base)

    # Priority 2: zero-stripped, full (strip 0610→610 before removing any segment)
    base = '-'.join(segs_stripped)
    if len(base) <= base_budget:
        return _assemble(base)

    # Priority 3: original with interior-segment reduction
    base = _try_fit(segs_orig, base_budget)
    if base is not None:
        return _assemble(base)

    # Priority 4: stripped with interior-segment reduction
    base = _try_fit(segs_stripped, base_budget)
    if base is not None:
        return _assemble(base)

    # --- Fallback: hard truncate ---
    base_raw = '-'.join(segs_orig)
    return (base_raw + suffix)[:18]


# Legacy 2-arg shim (backward compat — maps item_num to line_index on sheet 1)
def _build_old_material_number_legacy(spir_no: str, item_num: int) -> str:
    """Legacy 2-arg wrapper: sheet_index=1, line_index=item_num."""
    return build_old_material_number(spir_no, 1, item_num)


def build_position_number(item_num: int) -> str:
    """
    POSITION NUMBER = item_num * 10, zero-padded to 4 digits.
    item 1 → '0010', item 2 → '0020', item 10 → '0100'
    """
    return str(item_num * 10).zfill(4)


# ── Metadata reader ───────────────────────────────────────────────────────────

def _read_metadata(ws) -> dict:
    """
    Read equipment metadata from header rows (1-7).
    EQPT MAKE = MANUFACTURER (the equipment maker, from row 3 right-side).
    """
    meta = {
        "equipment_desc": "",
        "manufacturer":   "",   # → EQPT MAKE
        "supplier":       "",
        "project_name":   "",
        "spir_type":      _detect_spir_type(ws),
    }

    v = _v(ws, 2, 24)
    if v: meta["project_name"] = v
    v = _v(ws, 3, 25)
    if v: meta["manufacturer"] = v
    v = _v(ws, 4, 23)
    if v and not meta["supplier"]: meta["supplier"] = v

    for ri in range(2, 5):
        for ci in range(1, min(ws.max_column + 1, 30)):
            label = _v(ws, ri, ci).lower()
            if "manufacturer:" in label or label == "manufacturer:":
                for dc in range(1, 8):
                    v = _v(ws, ri, ci + dc)
                    if v and not any(kw in v.lower() for kw in
                                     ["manufacturer", "supplier", "equipment", "authority"]):
                        meta["manufacturer"] = meta["manufacturer"] or v
                        break
            if "supplier:" in label or label == "supplier:":
                for dc in range(1, 8):
                    v = _v(ws, ri, ci + dc)
                    if v and not any(kw in v.lower() for kw in
                                     ["supplier", "manufacturer", "equipment"]):
                        meta["supplier"] = meta["supplier"] or v
                        break
            if "equipment:" in label or label == "equipment:":
                for dc in range(1, 8):
                    v = _v(ws, ri, ci + dc)
                    if v and not any(kw in v.lower() for kw in ["equipment"]):
                        meta["project_name"] = meta["project_name"] or v
                        break

    return meta


# ── Column map builder ────────────────────────────────────────────────────────

def _build_col_map(ws, item_col: int) -> dict[int, str]:
    """Build data column → field name mapping from row 7 labels and row 6 headers."""
    col_map: dict[int, str] = {item_col: "item_num"}

    for c in range(item_col + 1, ws.max_column + 1):
        label7 = _v(ws, 7, c).lower().replace(" ", "")
        if label7 in _COL_LABEL_TO_FIELD:
            col_map[c] = _COL_LABEL_TO_FIELD[label7]
            continue
        label6 = _v(ws, 6, c).lower()
        for kw, field in _HEADER_FALLBACK.items():
            if kw in label6:
                col_map[c] = field
                break

    # Ensure unit_price is always mapped
    if not any(f == "unit_price" for f in col_map.values()):
        currency_col = None
        for c, f in col_map.items():
            if f == "currency":
                currency_col = c
                break
        if currency_col:
            col_map[currency_col + 1] = "unit_price"

    if not any(f == "qty_identical" for f in col_map.values()):
        col_map[item_col + 1] = "qty_identical"
    if not any(f == "desc" for f in col_map.values()):
        col_map[item_col + 2] = "desc"

    return col_map


# ── Item data reader ──────────────────────────────────────────────────────────

def _read_item_data(ws, r: int, col_map: dict[int, str],
                    sheet_name: str, spir_no: str, spir_type: str) -> dict:
    """Read one spare row into a dict using the column map."""
    data: dict = {
        "sheet":     sheet_name.strip().upper(),
        "spir_no":   spir_no,
        "spir_type": _normalise_spir_type(spir_type),
    }
    for c, field in col_map.items():
        if field == "item_num": continue
        val = _raw(ws, r, c)
        if val is None: continue
        s = str(val).strip()
        if not s or s == ".": continue
        if field == "qty_identical":
            try: val = int(float(s))
            except: pass
        elif field == "unit_price":
            try: val = float(s)
            except: pass
        data[field] = val

    SKIP = {"na", "tba", "n/a", "tbd", "-", "--", "none", "nil"}

    def _clean(v) -> str:
        if v is None: return ""
        s = str(v).strip()
        return "" if s.lower() in SKIP else s

    desc        = _clean(data.get("desc", ""))
    mfr_part    = _clean(data.get("mfr_part_no", ""))
    supp_part   = _clean(data.get("supplier_part_no", ""))
    supplier_nm = _clean(data.get("supplier_name", ""))

    if mfr_part:
        data["mfr_part_no"] = mfr_part
    elif supp_part:
        data["mfr_part_no"] = supp_part
    else:
        data.pop("mfr_part_no", None)

    final_mfr = data.get("mfr_part_no", "")

    if desc:
        parts = [desc]
        if final_mfr:   parts.append(str(final_mfr).strip())
        if supplier_nm: parts.append(supplier_nm)
        data["new_desc"] = ", ".join(parts)

    unit_price = data.get("unit_price")
    currency   = data.get("currency", "")
    if unit_price is not None and currency:
        try:
            from services.currency_service import to_qar
            qar = to_qar(float(unit_price), currency)
            if qar is not None:
                data["unit_price_qar"] = qar
        except Exception:
            pass

    return data


# ── Sheet classifier ──────────────────────────────────────────────────────────

def _classify_sheet(sn: str, ws) -> str:
    lower = sn.lower()
    if "validation" in lower: return "skip"

    has_item_header = any(
        "item number" in _v(ws, 6, c).lower()
        for c in range(1, ws.max_column + 1)
    )
    has_desc = any(
        "description" in _v(ws, r, c).lower()
        for r in [6, 7] for c in range(1, ws.max_column + 1)
    )

    if "continuation" not in lower and "annexure" not in lower:
        return "main"
    if "annexure" in lower:
        return "main"

    if "continuation" in lower:
        has_tags = any(
            _is_tag(_v(ws, 1, c))
            for c in range(4, min(ws.max_column + 1, 50))
        )
        if not has_tags:
            has_real_data = False
            for r in range(8, min(ws.max_row + 1, 20)):
                for c in range(4, min(ws.max_column + 1, 41)):
                    v = ws.cell(r, c).value
                    if v is not None and str(v).strip() not in ('', '0', '.', 'None'):
                        try:
                            if float(str(v)) > 0:
                                has_real_data = True
                                break
                        except (ValueError, TypeError):
                            has_real_data = True
                            break
                if has_real_data:
                    break
            if not has_real_data:
                return "skip"

    if has_desc:
        return "cont_full"
    if has_item_header:
        return "cont_marker"

    has_tags = any(
        _is_tag(_v(ws, 1, c))
        for c in range(3, min(ws.max_column + 1, 50))
    )
    return "cont_overflow_tagged" if has_tags else "cont_overflow_notag"


# ── Row builder ───────────────────────────────────────────────────────────────

def _build_output_row(row_data: dict, is_header: bool, spir_no: str,
                      item_num: Optional[int] = None,
                      sheet_index: int = 1,
                      line_index: int = 1) -> dict:
    """
    Build a standardised output row dict with all 27 columns.
    is_header=True  → summary row (EQPT_QTY filled, ITEM NUMBER empty)
    is_header=False → detail row (ITEM NUMBER + POSITION NUMBER + OLD MAT NUM filled)

    sheet_index : 1-based index of the source sheet (for OLD MATERIAL NUMBER)
    line_index  : 1-based line counter within this sheet (for OLD MATERIAL NUMBER)
    """
    out = {col: None for col in OUTPUT_COLS}

    out['SPIR NO']    = row_data.get('spir_no', spir_no)
    out['TAG NO']     = row_data.get('tag_no')
    out['EQPT MAKE']  = row_data.get('manufacturer', '')
    out['EQPT MODEL'] = row_data.get('model', '')
    out['EQPT SR NO'] = row_data.get('serial', '')
    out['SPIR ERROR'] = 0
    out['SHEET']      = row_data.get('sheet', '')
    out['SPIR TYPE']  = _normalise_spir_type(row_data.get('spir_type', ''))

    if is_header:
        out['EQPT QTY']             = row_data.get('eqpt_qty', 1)
        out['DESCRIPTION OF PARTS'] = row_data.get('desc', '')
        out['POSITION NUMBER']      = '0010'
    else:
        if item_num is not None:
            out['ITEM NUMBER']                    = item_num
            out['POSITION NUMBER']                = build_position_number(item_num)
            out['OLD MATERIAL NUMBER/SPF NUMBER'] = build_old_material_number(
                spir_no, sheet_index, line_index)
        out['QUANTITY IDENTICAL PARTS FITTED'] = row_data.get('qty_identical', 1)
        out['DESCRIPTION OF PARTS']            = row_data.get('desc', '')
        out['NEW DESCRIPTION OF PARTS']        = row_data.get('new_desc', '')
        out['DWG NO INCL POSN NO']             = row_data.get('dwg_no', '')
        out['MANUFACTURER PART NUMBER']        = row_data.get('mfr_part_no', '')
        out['MATERIAL SPECIFICATION']          = row_data.get('material_spec', '')
        out['SUPPLIER OCM NAME']               = row_data.get('supplier_name', '')
        out['CURRENCY']                        = row_data.get('currency', '')
        out['UNIT PRICE']                      = row_data.get('unit_price')
        out['UNIT PRICE (QAR)']                = row_data.get('unit_price_qar')
        out['DELIVERY TIME IN WEEKS']          = row_data.get('delivery')
        out['MIN MAX STOCK LVLS QTY']          = row_data.get('min_max')
        out['UNIT OF MEASURE']                 = row_data.get('uom', '')
        out['SAP NUMBER']                      = row_data.get('sap_no', '')
        out['CLASSIFICATION OF PARTS']         = row_data.get('classification', '')

    return out


# ── Main sheet extractor ──────────────────────────────────────────────────────

# ── SPIR FORMAT TYPE DETECTION ────────────────────────────────────────────────
# The system handles 8 distinct SPIR sheet formats.
# Format detection is based on structural signals (tag layout, sheet names, etc.)
# This drives separate per-format logic blocks — DO NOT merge these branches.

SPIR_FORMAT_UNKNOWN      = 0
SPIR_FORMAT_SINGLE_TAG   = 1   # One tag column per sheet
SPIR_FORMAT_MULTI_TAG    = 2   # Multiple single-value tag columns
SPIR_FORMAT_PACKED_TAGS  = 3   # Multiple tags packed in ONE cell (e.g. 8 GV tags)
SPIR_FORMAT_MULTI_SHEET  = 4   # Same tags repeated across 2+ main sheets
SPIR_FORMAT_CONTINUATION = 5   # Overflow/continuation sheet with packed tags
SPIR_FORMAT_ANNEXURE     = 6   # Annexure-style (tag rows in separate sheet)
SPIR_FORMAT_MIXED        = 7   # Mixed: some single-tag cols + packed-tag col(s)
SPIR_FORMAT_LEGACY       = 8   # Legacy formats 1-5 (handled by existing code)


def _detect_spir_format(ws, sheet_name: str, wb=None) -> int:
    """
    Detect the structural SPIR format for a given sheet.

    Signals examined:
    - Number of tag columns in R1
    - Whether any tag cell contains ',' or '/' (packed tags)
    - Whether multiple main sheets exist (wb provided)
    - Sheet name keywords (continuation, annexure)
    """
    lower = sheet_name.lower()

    # Format 6: Annexure style detected at workbook level (checked before calling)
    if 'annexure' in lower:
        return SPIR_FORMAT_ANNEXURE

    # Format 5: Continuation sheet
    if 'continuation' in lower:
        return SPIR_FORMAT_CONTINUATION

    # Scan R1 for tag columns (between col 3 and item_col)
    tag_values = []
    for c in range(3, min(ws.max_column + 1, 30)):
        v = _v(ws, 1, c)
        if _is_tag(v):
            tag_values.append(v)
        elif v and 'item number' in _v(ws, 6, c).lower():
            break

    if not tag_values:
        return SPIR_FORMAT_UNKNOWN

    packed_cols  = [v for v in tag_values if ',' in v or '/' in v or ' to ' in v.lower()]
    single_cols  = [v for v in tag_values if v not in packed_cols]

    # Format 3: ALL tag columns are packed (e.g. one cell = '30-GV-146, 171, 169...')
    if packed_cols and not single_cols:
        return SPIR_FORMAT_PACKED_TAGS

    # Format 7: Mixed — some single-tag cols + at least one packed-tag col
    if packed_cols and single_cols:
        return SPIR_FORMAT_MIXED

    # Format 4: Multi-sheet (same tags in 2+ sheets) — detected from wb
    if wb is not None:
        main_sheet_count = sum(
            1 for sn in wb.sheetnames
            if 'validation' not in sn.lower()
            and 'continuation' not in sn.lower()
            and 'annexure' not in sn.lower()
        )
        if main_sheet_count > 1:
            return SPIR_FORMAT_MULTI_SHEET

    # Format 2: Multiple single-value tag columns
    if len(single_cols) > 1:
        return SPIR_FORMAT_MULTI_TAG

    # Format 1: Single tag column
    return SPIR_FORMAT_SINGLE_TAG


def _distribute_qty_per_tag(total_identical: int, num_tags: int,
                             eqpt_qty_header: int | None) -> tuple[int, int]:
    """
    SPIR 6 special case: distribute identical parts and eqpt_qty per tag.

    Rules (per requirements):
    - 8 tags, eqpt_qty=EMPTY, identical=8  → each tag: eqpt_qty=1, identical=1
    - 10 tags, qty=10                       → each tag: eqpt_qty=1, identical=1
    - 10 tags, qty=20                       → each tag: eqpt_qty=2, identical=2

    Returns: (per_tag_eqpt_qty, per_tag_identical)
    """
    if num_tags <= 0:
        return (eqpt_qty_header or 1, total_identical)

    # Per-tag identical = total / num_tags (floor division, minimum 1)
    per_tag_identical = max(1, total_identical // num_tags)

    # Per-tag eqpt_qty:
    # If header eqpt_qty was None/empty → each tag gets 1
    # If header eqpt_qty was set → divide by num_tags
    if eqpt_qty_header is None or eqpt_qty_header == 0:
        per_tag_eqpt_qty = 1
    else:
        per_tag_eqpt_qty = max(1, eqpt_qty_header // num_tags)

    return (per_tag_eqpt_qty, per_tag_identical)


# ── Main sheet extractor (dispatches per SPIR format) ─────────────────────────

def _extract_main(ws, sheet_name: str, spir_no: str, wb=None):
    """
    Extract items, tag-item links, and tag metadata from one main sheet.

    Dispatches to separate logic blocks per SPIR format type.
    DO NOT merge these blocks — kept separate for registry extensibility.

    FORMAT DISPATCH:
      SPIR_FORMAT_SINGLE_TAG  (1) — one tag col, direct cell value → qty_identical
      SPIR_FORMAT_MULTI_TAG   (2) — multiple independent single-tag cols
      SPIR_FORMAT_PACKED_TAGS (3) — all tags packed in one cell (SPIR 6 GV col)
      SPIR_FORMAT_MULTI_SHEET (4) — same tag set repeated across 2+ main sheets
      SPIR_FORMAT_MIXED       (7) — single-tag cols + packed-tag col(s)

    LINE NUMBER SHARING RULE (CRITICAL):
      line_index = position of the item in its source sheet (1-based counter).
      ALL tags that reference the same item SHARE the same line_index.
      Never increment line_index per tag — only per item position.
    """
    if ws.max_row < 7:
        return {}, {}, {}

    # ── Detect item column ────────────────────────────────────────────────────
    item_col = None
    for c in range(1, ws.max_column + 1):
        if "item number" in _v(ws, 6, c).lower():
            item_col = c
            break

    if item_col is None:
        desc_col = None
        for c in range(1, ws.max_column + 1):
            if "description" in _v(ws, 7, c).lower():
                desc_col = c
                break
        if desc_col is not None:
            for candidate in [desc_col - 1, desc_col - 2]:
                if candidate < 1:
                    continue
                vals = [ws.cell(r, candidate).value
                        for r in range(8, min(ws.max_row + 1, 16))]
                nums = [int(float(str(v))) for v in vals
                        if v is not None and str(v).strip() not in ('', '.')]
                if (len(nums) >= 3 and
                        nums == list(range(nums[0], nums[0] + len(nums)))):
                    item_col = candidate
                    break

    if item_col is None:
        return {}, {}, {}

    # ── Scan R1 for tag columns ───────────────────────────────────────────────
    tags_raw, tag_cols = [], []
    for c in range(3, item_col):
        val = _v(ws, 1, c)
        if _is_tag(val):
            tags_raw.append(val)
            tag_cols.append(c)
    if not tags_raw:
        return {}, {}, {}

    # ── Classify each column as packed (multi-tag) or single ─────────────────
    packed_col_set: set[int] = set()
    for raw, col in zip(tags_raw, tag_cols):
        if ',' in raw or '/' in raw or re.search(r'\bto\b', raw, re.IGNORECASE):
            packed_col_set.add(col)

    # ── Detect SPIR format for this sheet ────────────────────────────────────
    spir_fmt = _detect_spir_format(ws, sheet_name, wb)

    # ── Read header rows ──────────────────────────────────────────────────────
    models    = [_v(ws, 4, c) for c in tag_cols]
    serials   = [_v(ws, 6, c) for c in tag_cols]

    # eqpt_qty: None when cell is empty (important for packed-col distribution)
    eqpt_qtys: list[int | None] = []
    for c in tag_cols:
        v = _raw(ws, 7, c)
        if v is None:
            eqpt_qtys.append(None)
        else:
            try:
                iq = int(float(str(v)))
                eqpt_qtys.append(iq if iq > 0 else None)
            except Exception:
                eqpt_qtys.append(None)

    col_map   = _build_col_map(ws, item_col)
    spir_type = _detect_spir_type(ws)
    meta      = _read_metadata(ws)

    # ── SPIR type override columns ────────────────────────────────────────────
    spir_type_cols: dict[int, str] = {}
    for c in range(item_col, min(ws.max_column + 1, item_col + 20)):
        for r in [5, 6]:
            v = _v(ws, r, c).lower()
            if "initial spare" in v:
                spir_type_cols[c] = "Initial Spare Parts"
            elif "normal" in v and "spare" in v:
                spir_type_cols[c] = "Normal Operating Spares"
            elif "commissioning" in v:
                spir_type_cols[c] = "Commissioning Spare Parts"
            elif "life cycle" in v:
                spir_type_cols[c] = "Life Cycle Spare Parts"

    # ── Build tag_meta with per-format eqpt_qty distribution ─────────────────
    tag_meta: dict[str, dict] = {}

    if spir_fmt in (SPIR_FORMAT_SINGLE_TAG,
                    SPIR_FORMAT_MULTI_TAG,
                    SPIR_FORMAT_MULTI_SHEET):
        # FORMAT 1 / 2 / 4: single-value tag columns
        # Each tag column is independent — use header eqpt_qty directly (default 1)
        for i, tag_raw in enumerate(tags_raw):
            for tag in _expand_tags(tag_raw):
                if not tag:
                    continue
                hdr_qty = eqpt_qtys[i] if i < len(eqpt_qtys) else None
                tag_meta[tag] = {
                    "model":        models[i]  if i < len(models)  else "",
                    "serial":       serials[i] if i < len(serials) else "",
                    "eqpt_qty":     hdr_qty if hdr_qty is not None else 1,
                    "num_tags_in_col": 1,
                    "sheet":        sheet_name.strip().upper(),
                    "manufacturer": meta["manufacturer"],
                    "project_name": meta["project_name"],
                }

    elif spir_fmt == SPIR_FORMAT_PACKED_TAGS:
        # FORMAT 3: all tags packed in one cell per column
        # SPIR 6 rule: eqpt_qty=EMPTY → each tag gets eqpt_qty=1
        for i, tag_raw in enumerate(tags_raw):
            expanded = _expand_tags(tag_raw)
            num_t    = max(1, len(expanded))
            hdr_qty  = eqpt_qtys[i] if i < len(eqpt_qtys) else None
            per_tag_eq, _ = _distribute_qty_per_tag(num_t, num_t, hdr_qty)
            for tag in expanded:
                if not tag:
                    continue
                tag_meta[tag] = {
                    "model":        models[i]  if i < len(models)  else "",
                    "serial":       serials[i] if i < len(serials) else "",
                    "eqpt_qty":     per_tag_eq,
                    "num_tags_in_col": num_t,
                    "sheet":        sheet_name.strip().upper(),
                    "manufacturer": meta["manufacturer"],
                    "project_name": meta["project_name"],
                }

    elif spir_fmt == SPIR_FORMAT_MIXED:
        # FORMAT 7: some cols are single-tag, others are packed
        # Packed cols use distribution; single cols use direct value
        for i, tag_raw in enumerate(tags_raw):
            col      = tag_cols[i]
            expanded = _expand_tags(tag_raw)
            is_pk    = col in packed_col_set
            num_t    = max(1, len(expanded)) if is_pk else 1
            hdr_qty  = eqpt_qtys[i] if i < len(eqpt_qtys) else None
            if is_pk:
                per_tag_eq, _ = _distribute_qty_per_tag(num_t, num_t, hdr_qty)
            else:
                per_tag_eq = hdr_qty if hdr_qty is not None else 1
            for tag in expanded:
                if not tag:
                    continue
                tag_meta[tag] = {
                    "model":        models[i]  if i < len(models)  else "",
                    "serial":       serials[i] if i < len(serials) else "",
                    "eqpt_qty":     per_tag_eq,
                    "num_tags_in_col": num_t,
                    "sheet":        sheet_name.strip().upper(),
                    "manufacturer": meta["manufacturer"],
                    "project_name": meta["project_name"],
                }

    else:
        # FALLBACK: treat every column as independent (FORMAT 2 path)
        for i, tag_raw in enumerate(tags_raw):
            for tag in _expand_tags(tag_raw):
                if not tag:
                    continue
                hdr_qty = eqpt_qtys[i] if i < len(eqpt_qtys) else None
                tag_meta[tag] = {
                    "model":        models[i]  if i < len(models)  else "",
                    "serial":       serials[i] if i < len(serials) else "",
                    "eqpt_qty":     hdr_qty if hdr_qty is not None else 1,
                    "num_tags_in_col": 1,
                    "sheet":        sheet_name.strip().upper(),
                    "manufacturer": meta["manufacturer"],
                    "project_name": meta["project_name"],
                }

    # ── Read data rows ────────────────────────────────────────────────────────
    items:     dict[tuple, dict]      = {}
    tag_items: dict[str, list[tuple]] = {}

    for r in range(8, ws.max_row + 1):
        item_val = _raw(ws, r, item_col)
        if item_val is None:
            continue
        try:
            item_num = int(float(str(item_val)))
        except Exception:
            continue
        if item_num <= 0:
            continue

        key = (sheet_name, item_num)

        # Determine row-level SPIR type (initial/normal/commissioning/lifecycle)
        row_spir_type = spir_type
        for tc, type_label in spir_type_cols.items():
            cell_val = _raw(ws, r, tc)
            if cell_val is not None and cell_val not in (0, "0", "", "."):
                try:
                    if float(str(cell_val)) > 0:
                        row_spir_type = type_label
                        break
                except (ValueError, TypeError):
                    pass

        data = _read_item_data(ws, r, col_map, sheet_name, spir_no, row_spir_type)
        if data.get("desc") or data.get("mfr_part_no"):
            items[key] = data

        # ── Tag-item linkage with per-format qty_identical distribution ───────
        # LINE NUMBER SHARING: line_index is set by item position (not per tag).
        # All tags linked to the same item share the same line_index in
        # build_old_material_number — enforced in Pass 1 of extract_workbook.

        for i, (tag_raw, tc) in enumerate(zip(tags_raw, tag_cols)):
            cell_val = _raw(ws, r, tc)
            if cell_val is None:
                continue

            # Parse the flag/qty value from the cell
            try:
                cell_num = float(str(cell_val))
                if cell_num <= 0:
                    continue
            except (ValueError, TypeError):
                s_val = str(cell_val).strip()
                if not s_val or s_val in ("0", "."):
                    continue
                cell_num = 1.0   # non-numeric non-empty → treat as 1

            expanded_tags = _expand_tags(tag_raw)
            if not expanded_tags:
                continue

            is_packed = tc in packed_col_set
            num_t     = max(1, len(expanded_tags)) if is_packed else 1

            # FORMAT 1: single tag column
            if spir_fmt == SPIR_FORMAT_SINGLE_TAG:
                per_tag_identical = int(cell_num)

            # FORMAT 2 / 4: multiple independent single-tag cols
            elif spir_fmt in (SPIR_FORMAT_MULTI_TAG, SPIR_FORMAT_MULTI_SHEET):
                per_tag_identical = int(cell_num)

            # FORMAT 3: all cols are packed — distribute total across tags
            elif spir_fmt == SPIR_FORMAT_PACKED_TAGS:
                # cell_num = total qty for all tags in this packed col
                # SPIR 6 rule: per_tag = total / num_tags (floor, min 1)
                _, per_tag_identical = _distribute_qty_per_tag(
                    int(cell_num), num_t,
                    eqpt_qtys[i] if i < len(eqpt_qtys) else None)

            # FORMAT 7: mixed — packed cols distribute, single cols are direct
            elif spir_fmt == SPIR_FORMAT_MIXED:
                if is_packed:
                    _, per_tag_identical = _distribute_qty_per_tag(
                        int(cell_num), num_t,
                        eqpt_qtys[i] if i < len(eqpt_qtys) else None)
                else:
                    per_tag_identical = int(cell_num)

            else:
                # FALLBACK
                per_tag_identical = max(1, int(cell_num) // max(1, num_t))

            for tag in expanded_tags:
                if not tag:
                    continue
                # KEY FORMAT: always 3-tuple (sheet, item_num, tc)
                # This ensures ts_col is preserved for ALL formats (not just packed).
                # Without tc in the key, the output builder's effective_tc is None
                # and qty_by_col lookup falls back to the wrong total-identical value.
                # _get_item_data normalises to 2-tuple when looking up master_items,
                # so this change is safe for all downstream consumers.
                link_key = (sheet_name, item_num, tc)
                tag_items.setdefault(tag, []).append(link_key)
                # Store per-col qty so output builder uses the tag's own value,
                # not the total-across-all-tags value from the qty_identical column.
                if key in items:
                    items[key].setdefault('qty_by_col', {})[tc] = per_tag_identical

    return items, tag_items, tag_meta


# ── Continuation sheet extractor ──────────────────────────────────────────────

def _extract_cont_marker(ws, sheet_name: str, main_sheets: list[str],
                          master_items: dict) -> dict:
    tags_raw, tag_cols = [], []
    tag_models:     dict[int, str] = {}
    tag_serials:    dict[int, str] = {}
    tag_eqpt_qtys:  dict[int, int] = {}

    for c in range(4, ws.max_column + 1):
        val = _v(ws, 1, c)
        if _is_tag(val):
            tags_raw.append(val)
            tag_cols.append(c)
            tag_models[c]  = _v(ws, 4, c)
            tag_serials[c] = _v(ws, 6, c)
            try:
                eq = int(float(str(_raw(ws, 7, c) or 1)))
                tag_eqpt_qtys[c] = eq if eq > 0 else 1
            except:
                tag_eqpt_qtys[c] = 1

    if not tags_raw: return {}

    cont_sheet_upper = sheet_name.strip().upper()
    tag_items:    dict[str, list[tuple]] = {}
    tag_meta_out: dict[str, dict]        = {}

    for r in range(8, ws.max_row + 1):
        item_val = _raw(ws, r, 3)
        if item_val is None: continue
        try: item_num = int(float(str(item_val)))
        except: continue
        if item_num <= 0: continue

        item_exists = any((ms, item_num) in master_items for ms in main_sheets)
        if not item_exists: continue

        for tag_raw, tc in zip(tags_raw, tag_cols):
            cell_val = _raw(ws, r, tc)
            if cell_val is None: continue
            try:
                if int(float(str(cell_val))) <= 0: continue
            except:
                s = str(cell_val).strip()
                if not s or s in ("0", "."): continue

            for tag in _expand_tags(tag_raw):
                if not tag: continue
                key = (cont_sheet_upper, item_num, tc)
                tag_items.setdefault(tag, []).append(key)
                meta_key = (tag, tc)
                if meta_key not in tag_meta_out:
                    tag_meta_out[meta_key] = {
                        "model":        tag_models.get(tc, ""),
                        "serial":       tag_serials.get(tc, ""),
                        "eqpt_qty":     tag_eqpt_qtys.get(tc, 1),
                        "sheet":        cont_sheet_upper,
                        "manufacturer": "",
                        "project_name": "",
                    }

    return {"tag_items": tag_items, "tag_meta": tag_meta_out}


def _extract_cont_overflow_tagged(ws, sheet_name: str, main_sheets: list[str],
                                   master_items: dict) -> tuple[dict, dict]:
    tags_raw: list[str] = []
    for c in range(3, min(ws.max_column + 1, 15)):
        val = _v(ws, 1, c)
        if _is_tag(val):
            tags_raw.append(val)

    if not tags_raw: return {}, {}

    item_nums: list[int] = []
    for r in range(8, ws.max_row + 1):
        v = _raw(ws, r, 4)
        if v is None: continue
        try:
            n = int(float(str(v)))
            if n > 0: item_nums.append(n)
        except: pass

    tag_items: dict[str, list[tuple]] = {}
    tag_meta:  dict[str, dict]        = {}
    for tag_raw in tags_raw:
        for tag in _expand_tags(tag_raw):
            if not tag: continue
            tag_meta[tag] = {"model": "", "serial": "", "eqpt_qty": 1,
                             "sheet": sheet_name.strip().upper(),
                             "manufacturer": "", "project_name": ""}
            for item_num in item_nums:
                for msheet in main_sheets:
                    key = (msheet, item_num)
                    if key in master_items:
                        tag_items.setdefault(tag, []).append(key)
                        break
    return tag_items, tag_meta


def _extract_cont_overflow_notag(ws, main_sheets: list[str],
                                  master_items: dict,
                                  existing_tags: dict) -> dict[str, list[tuple]]:
    """
    Handles continuation sheets that have no explicit tag headers but list
    item numbers in col 3. Only adds items genuinely in master_items.
    """
    item_nums: list[int] = []
    for r in range(8, ws.max_row + 1):
        v = _raw(ws, r, 3)
        if v is None: continue
        try:
            n = int(float(str(v)))
            if n > 0: item_nums.append(n)
        except: pass

    any_real_item = any(
        (ms, n) in master_items
        for n in item_nums
        for ms in main_sheets
    )
    if not any_real_item:
        return {}

    additions: dict[str, list[tuple]] = {}
    for tag in existing_tags:
        existing_keys = set(existing_tags.get(tag, []))
        for item_num in item_nums:
            for msheet in main_sheets:
                key = (msheet, item_num)
                if key in master_items and key not in existing_keys:
                    additions.setdefault(tag, []).append(key)
                    break
    return additions


# ── Annexure-style extractor ──────────────────────────────────────────────────

def _is_annexure_style(wb) -> bool:
    for sn in wb.sheetnames:
        if "validation" in sn.lower() or "continuation" in sn.lower(): continue
        ws = wb[sn]
        for c in range(1, min(ws.max_column + 1, 12)):
            if "annexure" in _v(ws, 1, c).lower():
                return True
    return False


def _extract_annexure_style(wb, spir_no: str) -> dict:
    """
    Annexure format: single MAIN SHEET + Annexure sheets.
    Tags live in Annexure sheets (rows), not as column headers.
    """
    main_ws, main_name = None, ""
    for sn in wb.sheetnames:
        if "validation" in sn.lower() or "annexure" in sn.lower(): continue
        main_ws = wb[sn]; main_name = sn; break
    if main_ws is None:
        return _empty_result(spir_no)

    item_col = None
    for c in range(1, main_ws.max_column + 1):
        if "item number" in _v(main_ws, 6, c).lower():
            item_col = c; break
    if item_col is None: item_col = 9

    annexure_cols: list[tuple[int, str]] = []
    for c in range(1, item_col):
        v = _v(main_ws, 1, c)
        if "annexure" in v.lower():
            annexure_cols.append((c, v.strip()))

    col_map      = _build_col_map(main_ws, item_col)
    spir_type    = _detect_spir_type(main_ws)
    meta         = _read_metadata(main_ws)
    manufacturer = meta["manufacturer"]
    sheet_idx    = _get_sheet_index(main_name)

    items: dict[int, dict] = {}
    item_to_annexures: dict[int, list[str]] = {}
    line_counter = 0

    for r in range(8, main_ws.max_row + 1):
        item_val = _raw(main_ws, r, item_col)
        if item_val is None: continue
        try: item_num = int(float(str(item_val)))
        except: continue
        if item_num <= 0: continue

        line_counter += 1
        data = _read_item_data(main_ws, r, col_map, main_name, spir_no, spir_type)
        if not data.get("desc"): continue
        if not data.get("manufacturer"):
            data["manufacturer"] = manufacturer
        data["_line_index"] = line_counter
        items[item_num] = data
        item_to_annexures[item_num] = []
        for ac, ann_name in annexure_cols:
            cv_val = _raw(main_ws, r, ac)
            if cv_val is not None and str(cv_val).strip() not in ("", "0", "."):
                item_to_annexures[item_num].append(ann_name)

    annexure_tags: dict[str, list[dict]] = {}
    for sn in wb.sheetnames:
        if "annexure" not in sn.lower(): continue
        ws = wb[sn]
        tags = []
        hdr_row = None
        for ri in range(1, min(ws.max_row + 1, 10)):
            s = '|'.join(str(ws.cell(ri, ci).value or '') for ci in range(1, 10)).lower()
            if 'sr. no' in s or 'sr.no' in s:
                hdr_row = ri; break
        if hdr_row:
            for ri in range(hdr_row + 2, ws.max_row + 1):
                tag    = _v(ws, ri, 5)
                serial = _v(ws, ri, 23)
                model  = _v(ws, ri, 24)
                if tag and _is_tag(tag):
                    tags.append({"tag": tag, "serial": serial, "model": model})
        else:
            for r in range(5, ws.max_row + 1):
                tag = _v(ws, r, 5)
                if _is_tag(tag):
                    tags.append({"tag": tag, "serial": _v(ws, r, 23), "model": _v(ws, r, 24)})
        if tags: annexure_tags[sn] = tags

    output_rows: list[dict] = []
    spare_count = 0
    all_tags: set[str] = set()

    for ac, ann_name in annexure_cols:
        ann_ws_name = next((k for k in annexure_tags
                            if ann_name.lower() in k.lower()), None)
        if not ann_ws_name:
            continue
        ann_items = [n for n, anns in item_to_annexures.items() if ann_name in anns]

        eqpt_qty_val = _raw(main_ws, 7, ac)
        try:
            eqpt_qty = int(float(str(eqpt_qty_val))) if eqpt_qty_val else len(annexure_tags[ann_ws_name])
        except:
            eqpt_qty = len(annexure_tags[ann_ws_name])

        # Global rows (no TAG NO)
        for item_num in ann_items:
            d = items.get(item_num, {})
            if not d: continue
            line_idx = d.get("_line_index", item_num)

            global_hdr = {col: None for col in OUTPUT_COLS}
            global_hdr['SPIR NO']              = spir_no
            global_hdr['EQPT MAKE']            = manufacturer
            global_hdr['EQPT QTY']             = eqpt_qty
            global_hdr['POSITION NUMBER']      = build_position_number(item_num)
            global_hdr['DESCRIPTION OF PARTS'] = meta.get('project_name', '')
            global_hdr['SPIR ERROR']           = 0
            global_hdr['SHEET']                = main_name.strip().upper()
            global_hdr['SPIR TYPE']            = _normalise_spir_type(spir_type)
            output_rows.append(global_hdr)

            global_det = {col: None for col in OUTPUT_COLS}
            global_det['SPIR NO']                         = spir_no
            global_det['EQPT MAKE']                       = manufacturer
            global_det['QUANTITY IDENTICAL PARTS FITTED'] = d.get('qty_identical', 1)
            global_det['ITEM NUMBER']                     = item_num
            global_det['POSITION NUMBER']                 = build_position_number(item_num)
            global_det['OLD MATERIAL NUMBER/SPF NUMBER']  = build_old_material_number(
                spir_no, sheet_idx, line_idx)
            global_det['DESCRIPTION OF PARTS']            = d.get('desc', '')
            global_det['NEW DESCRIPTION OF PARTS']        = d.get('new_desc', '')
            global_det['DWG NO INCL POSN NO']             = d.get('dwg_no', '')
            global_det['MANUFACTURER PART NUMBER']        = d.get('mfr_part_no', '')
            global_det['MATERIAL SPECIFICATION']          = d.get('material_spec', '')
            global_det['SUPPLIER OCM NAME']               = d.get('supplier_name', '')
            global_det['CURRENCY']                        = d.get('currency', '')
            global_det['UNIT PRICE']                      = d.get('unit_price')
            global_det['UNIT PRICE (QAR)']                = d.get('unit_price_qar')
            global_det['DELIVERY TIME IN WEEKS']          = d.get('delivery')
            global_det['MIN MAX STOCK LVLS QTY']          = d.get('min_max')
            global_det['UNIT OF MEASURE']                 = d.get('uom', '')
            global_det['SAP NUMBER']                      = d.get('sap_no', '')
            global_det['CLASSIFICATION OF PARTS']         = d.get('classification', '')
            global_det['SPIR ERROR']                      = 0
            global_det['SHEET']                           = d.get('sheet', main_name).strip().upper()
            global_det['SPIR TYPE']                       = _normalise_spir_type(
                d.get('spir_type', spir_type))
            output_rows.append(global_det)

        # Per-tag rows
        for ti in annexure_tags[ann_ws_name]:
            tag = ti["tag"]
            all_tags.add(tag)

            header_row = {col: None for col in OUTPUT_COLS}
            header_row['SPIR NO']              = spir_no
            header_row['TAG NO']               = tag
            header_row['EQPT MAKE']            = manufacturer
            header_row['EQPT MODEL']           = ti.get('model', '')
            header_row['EQPT SR NO']           = ti.get('serial', '')
            header_row['EQPT QTY']             = eqpt_qty
            header_row['POSITION NUMBER']      = '0010'
            header_row['DESCRIPTION OF PARTS'] = meta.get('project_name', '')
            header_row['SPIR ERROR']           = 0
            header_row['SHEET']                = main_name.strip().upper()
            header_row['SPIR TYPE']            = _normalise_spir_type(spir_type)
            output_rows.append(header_row)

            for item_num in ann_items:
                d = items.get(item_num, {})
                if not d: continue
                line_idx = d.get("_line_index", item_num)

                detail_row = {col: None for col in OUTPUT_COLS}
                detail_row['SPIR NO']                         = spir_no
                detail_row['TAG NO']                          = tag
                detail_row['EQPT MAKE']                       = manufacturer
                detail_row['EQPT MODEL']                      = ti.get('model', '')
                detail_row['EQPT SR NO']                      = ti.get('serial', '')
                detail_row['QUANTITY IDENTICAL PARTS FITTED'] = d.get('qty_identical', 1)
                detail_row['ITEM NUMBER']                     = item_num
                detail_row['POSITION NUMBER']                 = build_position_number(item_num)
                detail_row['OLD MATERIAL NUMBER/SPF NUMBER']  = build_old_material_number(
                    spir_no, sheet_idx, line_idx)
                detail_row['DESCRIPTION OF PARTS']            = d.get('desc', '')
                detail_row['NEW DESCRIPTION OF PARTS']        = d.get('new_desc', '')
                detail_row['DWG NO INCL POSN NO']             = d.get('dwg_no', '')
                detail_row['MANUFACTURER PART NUMBER']        = d.get('mfr_part_no', '')
                detail_row['MATERIAL SPECIFICATION']          = d.get('material_spec', '')
                detail_row['SUPPLIER OCM NAME']               = d.get('supplier_name', '')
                detail_row['CURRENCY']                        = d.get('currency', '')
                detail_row['UNIT PRICE']                      = d.get('unit_price')
                detail_row['UNIT PRICE (QAR)']                = d.get('unit_price_qar')
                detail_row['DELIVERY TIME IN WEEKS']          = d.get('delivery')
                detail_row['MIN MAX STOCK LVLS QTY']          = d.get('min_max')
                detail_row['UNIT OF MEASURE']                 = d.get('uom', '')
                detail_row['SAP NUMBER']                      = d.get('sap_no', '')
                detail_row['CLASSIFICATION OF PARTS']         = d.get('classification', '')
                detail_row['SPIR ERROR']                      = 0
                detail_row['SHEET']                           = d.get('sheet', main_name).strip().upper()
                detail_row['SPIR TYPE']                       = _normalise_spir_type(
                    d.get('spir_type', spir_type))
                output_rows.append(detail_row)
                spare_count += 1

    return {
        "format":         "SPIR_ANNEXURE",
        "spir_no":        spir_no,
        "equipment":      meta["project_name"],
        "manufacturer":   manufacturer,
        "supplier":       meta.get("supplier", ""),
        "spir_type":      _normalise_spir_type(spir_type),
        "eqpt_qty":       len(all_tags),
        "spare_items":    spare_count,
        "total_tags":     len(all_tags),
        "annexure_count": len(annexure_tags),
        "annexure_stats": {k: len(v) for k, v in annexure_tags.items()},
        "rows":           output_rows,
        "output_cols":    OUTPUT_COLS,
    }


# ── Empty result ──────────────────────────────────────────────────────────────

def _empty_result(spir_no: str) -> dict:
    return {
        "format": "UNKNOWN", "rows": [], "total_tags": 0, "spare_items": 0,
        "spir_no": spir_no, "eqpt_qty": 0, "annexure_count": 0,
        "annexure_stats": {}, "manufacturer": "", "supplier": "",
        "spir_type": "", "equipment": "", "output_cols": OUTPUT_COLS,
    }


# ── Main entry point ──────────────────────────────────────────────────────────

def extract_workbook(wb, spir_no: str = "", spir_filename: str = "") -> dict:
    """
    Main entry point. Auto-detects format and extracts all SPIR data.
    Returns dict with 'rows' (list of 27-column dicts) and metadata.
    """
    # Reset sheet index registry for each new workbook
    _reset_sheet_registry()

    if not spir_no:
        spir_no = _detect_spir_no(wb)
    if not spir_no and spir_filename:
        m = re.search(
            r'(VEN-[A-Z0-9]+-[A-Z0-9]+-\d+-\d+-[A-Z0-9.\-]+|'
            r'\d{4}-[A-Z]{2}-\d{2}-\d{2}-\d{2}-\d{3}[^\s_]*)',
            spir_filename, re.IGNORECASE)
        if m: spir_no = m.group(1)

    if _is_annexure_style(wb):
        log.info("SPIR format: ANNEXURE-STYLE")
        return _extract_annexure_style(wb, spir_no)

    log.info("extract_workbook: spir_no=%r  sheets=%s", spir_no, wb.sheetnames)

    master_items:  dict[tuple, dict]      = {}
    all_tag_items: dict[str, list[tuple]] = {}
    all_tag_meta:  dict[str, dict]        = {}
    main_sheets:   list[str]              = []
    annexure_stats: dict[str, int]        = {}
    manufacturer = supplier = spir_type = equipment = ""

    # Per-sheet line counters for OLD MATERIAL NUMBER (line_index = position within sheet)
    sheet_line_counters: dict[str, int] = {}  # sheet_name → next line_index

    # Pass 1: main / full sheets
    for sheet_name in wb.sheetnames:
        ws       = wb[sheet_name]
        category = _classify_sheet(sheet_name, ws)
        if category in ("skip", "cont_marker",
                        "cont_overflow_tagged", "cont_overflow_notag"):
            continue

        items, tag_items, tag_meta = _extract_main(ws, sheet_name, spir_no, wb=wb)
        if not items: continue

        # Register sheet index and assign line indices to each item
        sidx = _get_sheet_index(sheet_name)
        line_counter = 0
        for key in sorted(items.keys(), key=lambda k: k[1]):   # sort by item_num
            line_counter += 1
            items[key]["_sheet_index"] = sidx
            items[key]["_line_index"]  = line_counter
        sheet_line_counters[sheet_name] = line_counter

        master_items.update(items)
        main_sheets.append(sheet_name)

        for tag, keys in tag_items.items():
            all_tag_items.setdefault(tag, []).extend(keys)
            if tag not in all_tag_meta:
                all_tag_meta[tag] = tag_meta.get(tag, {})

        n = sum(len(v) for v in tag_items.values())
        if n: annexure_stats[sheet_name] = n

        if not manufacturer:
            meta = _read_metadata(ws)
            manufacturer = meta.get("manufacturer", "")
            equipment    = meta.get("project_name", "")
        if not spir_type:
            for d in items.values():
                t = d.get("spir_type", "")
                if t: spir_type = t; break

    # Pass 2: continuation sheets
    for sheet_name in wb.sheetnames:
        ws       = wb[sheet_name]
        category = _classify_sheet(sheet_name, ws)

        if category == "cont_marker":
            cont_result = _extract_cont_marker(ws, sheet_name, main_sheets, master_items)
            tag_items_c = cont_result.get("tag_items", {})
            tag_meta_c  = cont_result.get("tag_meta", {})

            for tag, keys in tag_items_c.items():
                all_tag_items.setdefault(tag, []).extend(keys)
                for k in keys:
                    col = k[2] if len(k) > 2 else None
                    mk  = (tag, col)
                    if mk not in all_tag_meta:
                        me = dict(tag_meta_c.get(mk, tag_meta_c.get(tag, {})))
                        me["manufacturer"] = me.get("manufacturer") or manufacturer
                        all_tag_meta[mk] = me
                if tag not in all_tag_meta:
                    me = dict(tag_meta_c.get(tag, {}))
                    me["manufacturer"] = me.get("manufacturer") or manufacturer
                    all_tag_meta[tag] = me

            n = sum(len(v) for v in tag_items_c.values())
            if n: annexure_stats[sheet_name] = n

        elif category == "cont_overflow_tagged":
            additions, tag_meta = _extract_cont_overflow_tagged(
                ws, sheet_name, main_sheets, master_items)
            for tag, keys in additions.items():
                all_tag_items.setdefault(tag, []).extend(keys)
                if tag not in all_tag_meta:
                    all_tag_meta[tag] = tag_meta.get(tag, {})
            n = sum(len(v) for v in additions.values())
            if n: annexure_stats[sheet_name] = n

        elif category == "cont_overflow_notag":
            additions = _extract_cont_overflow_notag(
                ws, main_sheets, master_items, all_tag_items)
            for tag, keys in additions.items():
                all_tag_items[tag] = all_tag_items.get(tag, []) + keys

    # Build output rows
    from collections import defaultdict
    tag_sheet_items: dict[tuple, list[tuple]] = defaultdict(list)
    tag_sheet_order: list[tuple] = []

    for tag, keys in all_tag_items.items():
        seen: set = set()
        for key in keys:
            if key not in seen:
                seen.add(key)
                ts = (tag, key[0], key[2]) if len(key) > 2 else (tag, key[0])
                tag_sheet_items[ts].append(key)
                if ts not in tag_sheet_order:
                    tag_sheet_order.append(ts)

    output_rows: list[dict] = []
    spare_count = 0

    def _get_item_data(key: tuple) -> Optional[dict]:
        # key may be 2-tuple (sheet, item_num) or 3-tuple (sheet, item_num, tc)
        # master_items only stores 2-tuple keys → always look up as (sheet, item_num)
        sheet_k  = key[0]
        item_num = key[1]
        lookup   = (sheet_k, item_num)
        if lookup in master_items:
            return master_items[lookup]
        for ms in main_sheets:
            alt = (ms, item_num)
            if alt in master_items:
                return master_items[alt]
        return None

    for ts_key in tag_sheet_order:
        tag   = ts_key[0]
        sheet = ts_key[1]
        keys  = tag_sheet_items[ts_key]
        if not keys: continue

        ts_col = ts_key[2] if len(ts_key) > 2 else None
        meta   = all_tag_meta.get((tag, ts_col)) or all_tag_meta.get(tag, {})
        fi     = _get_item_data(keys[0]) or {}

        output_sheet = sheet.strip().upper()
        eqpt_qty     = meta.get("eqpt_qty", 1)

        # ── Header row ────────────────────────────────────────────────────────
        header_row = {col: None for col in OUTPUT_COLS}
        header_row['SPIR NO']              = spir_no
        header_row['TAG NO']               = tag
        header_row['EQPT MAKE']            = meta.get("manufacturer", manufacturer)
        header_row['EQPT MODEL']           = meta.get("model", "")
        header_row['EQPT SR NO']           = meta.get("serial", "")
        header_row['EQPT QTY']             = eqpt_qty
        header_row['POSITION NUMBER']      = '0010'
        header_row['DESCRIPTION OF PARTS'] = equipment or fi.get("desc", "")
        header_row['SPIR ERROR']           = 0
        header_row['SHEET']                = output_sheet
        header_row['SPIR TYPE']            = _normalise_spir_type(
            fi.get("spir_type", spir_type))
        output_rows.append(header_row)

        # ── Detail rows ───────────────────────────────────────────────────────
        for key in keys:
            d = _get_item_data(key)
            if d is None: continue
            item_num  = key[1]
            # For packed-col tags: key is 3-tuple (sheet, item_num, tc)
            # tc used to look up per-col qty_identical override stored by _extract_main
            key_tc    = key[2] if len(key) > 2 else None
            sidx      = d.get("_sheet_index", 1)
            line_idx  = d.get("_line_index", item_num)

            # qty_identical: prefer per-col override (packed cols), else item-level value
            # Packed cols: total stored in qty_identical, per-tag in qty_by_col[tc]
            qty_by_col    = d.get("qty_by_col", {})
            # Use key_tc first (most specific: this exact packed col)
            # then ts_col (the group-level col from the ts_key)
            # then raw qty_identical (single-tag or unresolved)
            effective_tc  = key_tc if key_tc is not None else ts_col
            qty_identical = (qty_by_col[effective_tc]
                             if effective_tc is not None and effective_tc in qty_by_col
                             else d.get("qty_identical", 1))

            detail_row = {col: None for col in OUTPUT_COLS}
            detail_row['SPIR NO']                         = spir_no
            detail_row['TAG NO']                          = tag
            detail_row['EQPT MAKE']                       = meta.get("manufacturer", manufacturer)
            detail_row['EQPT MODEL']                      = meta.get("model", "")
            detail_row['EQPT SR NO']                      = meta.get("serial", "")
            detail_row['QUANTITY IDENTICAL PARTS FITTED'] = qty_identical
            detail_row['ITEM NUMBER']                     = item_num
            detail_row['POSITION NUMBER']                 = build_position_number(item_num)
            detail_row['OLD MATERIAL NUMBER/SPF NUMBER']  = build_old_material_number(
                spir_no, sidx, line_idx)
            detail_row['DESCRIPTION OF PARTS']            = d.get("desc", "")
            detail_row['NEW DESCRIPTION OF PARTS']        = d.get("new_desc", "")
            detail_row['DWG NO INCL POSN NO']             = d.get("dwg_no", "")
            detail_row['MANUFACTURER PART NUMBER']        = d.get("mfr_part_no", "")
            detail_row['MATERIAL SPECIFICATION']          = d.get("material_spec", "")
            detail_row['SUPPLIER OCM NAME']               = d.get("supplier_name", "")
            detail_row['CURRENCY']                        = d.get("currency", "")
            detail_row['UNIT PRICE']                      = d.get("unit_price")
            detail_row['UNIT PRICE (QAR)']                = d.get("unit_price_qar")
            detail_row['DELIVERY TIME IN WEEKS']          = d.get("delivery")
            detail_row['MIN MAX STOCK LVLS QTY']          = d.get("min_max")
            detail_row['UNIT OF MEASURE']                 = d.get("uom", "")
            detail_row['SAP NUMBER']                      = d.get("sap_no", "")
            detail_row['CLASSIFICATION OF PARTS']         = d.get("classification", "")
            detail_row['SPIR ERROR']                      = 0
            detail_row['SHEET']                           = output_sheet
            detail_row['SPIR TYPE']                       = _normalise_spir_type(
                d.get("spir_type", spir_type))
            output_rows.append(detail_row)
            spare_count += 1

    all_tags = {ts[0] for ts in tag_sheet_items}

    return {
        "format":         "SPIR_MATRIX",
        "spir_no":        spir_no,
        "equipment":      equipment,
        "manufacturer":   manufacturer,
        "supplier":       supplier,
        "spir_type":      _normalise_spir_type(spir_type),
        "eqpt_qty":       len(all_tags),
        "spare_items":    spare_count,
        "total_tags":     len(all_tags),
        "annexure_count": max(0, len(annexure_stats) - 1),
        "annexure_stats": annexure_stats,
        "rows":           output_rows,
        "output_cols":    OUTPUT_COLS,
    }