"""
extraction/spir_detector.py
────────────────────────────
File validation and format detection.

Responsibilities:
  1. validate_file()  — extension, size, openability checks
  2. detect_format()  — try all registered format parsers, return best match
  3. detect()         — richer response dict for the /inspect endpoint

Works standalone without spir_engine.py — uses the format parsers directly.
"""
from __future__ import annotations
import io
import logging
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

log = logging.getLogger(__name__)

_ALLOWED_EXT = {".xlsx", ".xlsm", ".xls", ".csv"}


class ValidationError(Exception):
    """Raised when the uploaded file fails pre-processing validation."""


# ─────────────────────────────────────────────────────────────────────────────
# File validation
# ─────────────────────────────────────────────────────────────────────────────

def validate_file(filename: str, content: bytes, max_mb: int = 2048) -> None:
    """
    Validate before touching any parsing logic.

    Raises:
        ValidationError with a human-readable message.
    """
    lower = (filename or "").lower()
    if not any(lower.endswith(ext) for ext in _ALLOWED_EXT):
        raise ValidationError(
            f"Unsupported file type '{filename}'. "
            f"Accepted: {', '.join(sorted(_ALLOWED_EXT))}"
        )

    size_mb = len(content) / 1_048_576
    if size_mb > max_mb:
        raise ValidationError(
            f"File too large ({size_mb:.1f} MB). Maximum: {max_mb} MB."
        )

    if not content:
        raise ValidationError("File is empty.")

    # Structural check
    try:
        if lower.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(content), nrows=1)
            if df.empty:
                raise ValidationError("CSV file contains no data rows.")
        else:
            wb = openpyxl.load_workbook(
                io.BytesIO(content), data_only=True, read_only=True
            )
            n = len(wb.sheetnames)
            wb.close()
            if n == 0:
                raise ValidationError("Workbook contains no sheets.")
    except ValidationError:
        raise
    except Exception as exc:
        raise ValidationError(f"Cannot open file: {exc}") from exc


# ─────────────────────────────────────────────────────────────────────────────
# Format detection
# ─────────────────────────────────────────────────────────────────────────────

def detect_format(content_or_wb, filename: str = "") -> str:
    """
    Detect the SPIR format of a file.

    Args:
        content_or_wb: raw bytes OR an already-open openpyxl Workbook.
        filename:      original filename (for extension hints).

    Returns:
        Format name string, e.g. "FORMAT6", "FORMAT_ADAPTIVE", "UNKNOWN".
    """
    # Try format parsers from the formats package
    try:
        from extraction.formats import get_all_parsers
        if isinstance(content_or_wb, (bytes, bytearray)):
            wb = openpyxl.load_workbook(
                io.BytesIO(content_or_wb), data_only=True, read_only=True
            )
            close_wb = True
        else:
            wb = content_or_wb
            close_wb = False

        sheets = wb.sheetnames
        for parser_cls in get_all_parsers():
            try:
                if parser_cls.detect(wb):
                    if close_wb:
                        wb.close()
                    return parser_cls.FORMAT_NAME
            except Exception:
                pass

        if close_wb:
            wb.close()

        # Heuristic sheet-name detection
        return _heuristic_detect(sheets)

    except Exception as exc:
        log.warning("Format detection error: %s", exc)
        return "UNKNOWN"


def _heuristic_detect(sheet_names: list[str]) -> str:
    """Fall-back heuristic based on sheet names alone."""
    names_lower = [s.lower() for s in sheet_names]

    # FORMAT8: "main sheet N (CATEGORY)"
    categorised = [n for n in names_lower
                   if "main sheet" in n and "(" in n]
    if len(categorised) >= 2:
        return "FORMAT8"

    # FORMAT7: "main sheet 1", "main sheet 2"
    numbered = [n for n in names_lower
                if "main sheet" in n and n.replace("main sheet", "").strip().isdigit()]
    if len(numbered) >= 2:
        return "FORMAT7"

    # FORMAT5: multiple continuation sheets
    cont = [n for n in names_lower if "continuation" in n]
    if len(cont) >= 2:
        return "FORMAT5"

    # FORMAT4/6: single continuation
    if len(cont) == 1:
        return "FORMAT6"

    # FORMAT1: annexure sheets
    if any("annexure" in n for n in names_lower):
        return "FORMAT1"

    if len(sheet_names) == 1:
        return "FORMAT2"

    return "FORMAT_ADAPTIVE"


def detect(content: bytes, filename: str = "") -> dict:
    """
    Full detection with rich metadata. Used by the /inspect endpoint.
    """
    lower = filename.lower()
    sheets: list[str] = []

    if lower.endswith(".csv"):
        fmt = "CSV"
        description = "CSV file"
    else:
        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(content), data_only=True, read_only=True
            )
            sheets = wb.sheetnames
            wb.close()
        except Exception as exc:
            return {"format": "UNKNOWN", "description": str(exc), "sheets": []}

        fmt = detect_format(content, filename)
        description = _FORMAT_DESCRIPTIONS.get(fmt, f"Format: {fmt}")

    return {"format": fmt, "description": description, "sheets": sheets}


_FORMAT_DESCRIPTIONS: dict[str, str] = {
    "FORMAT1":          "Multi-annexure SPIR",
    "FORMAT2":          "Single-sheet, single tag",
    "FORMAT3":          "Single-sheet, multiple tag columns",
    "FORMAT4":          "Matrix SPIR + single continuation sheet",
    "FORMAT5":          "Flag SPIR + multiple continuation sheets",
    "FORMAT6":          "Mixed single/multi-tag + single continuation",
    "FORMAT7":          "Multiple numbered main sheets",
    "FORMAT8":          "Categorised numbered main sheets",
    "FORMAT_ADAPTIVE":  "Unknown format — adaptive extraction",
    "CSV":              "CSV file",
}
